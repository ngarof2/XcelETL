from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                                 QLabel, QPushButton, QLineEdit, QTextEdit, QTreeWidget, 
                                 QTreeWidgetItem, QProgressBar, QFileDialog, QMessageBox,
                                 QTabWidget, QCheckBox, QComboBox, QFrame, QGroupBox, QHeaderView,
                                 QDialog, QDialogButtonBox, QSpinBox, QSplitter, QRadioButton)
from PySide6.QtCore import Qt, QThread, Signal, QTimer
from PySide6.QtGui import QColor, QFont
import os
import shutil
import json
import csv
import hashlib
import requests
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import threading
import queue
try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
try:
    from mutagen import File as MutagenFile
    from mutagen.id3 import ID3NoHeaderError
    MUTAGEN_AVAILABLE = True
except ImportError:
    MUTAGEN_AVAILABLE = False
import zipfile
import tarfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ScanWorker(QThread):
    """Worker thread for scanning files with signals for thread-safe GUI updates"""
    progress_updated = Signal(int)
    status_updated = Signal(str)
    scan_complete = Signal(int)
    
    def __init__(self, target_folder, known_types):
        super().__init__()
        self.target_folder = target_folder
        self.known_types = known_types
        self.cancel_flag = threading.Event()
        self.file_list = []
    
    def run(self):
        """Scan files in a separate thread"""
        self.status_updated.emit("Scanning files...")
        self.progress_updated.emit(0)
        
        for root, dirs, files in os.walk(self.target_folder):
            if self.cancel_flag.is_set():
                break
            
            for file in files:
                if self.cancel_flag.is_set():
                    break
                
                file_path = os.path.join(root, file)
                try:
                    stat = os.stat(file_path)
                    size_kb = stat.st_size / 1024
                    modified = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
                    ext = os.path.splitext(file)[1].lower().lstrip('.')
                    
                    # Auto-categorize based on extension
                    if ext in self.known_types:
                        category = self.known_types[ext]
                    else:
                        category = 'Unknown'
                    
                    self.file_list.append({
                        'filename': file,
                        'path': file_path,
                        'size_kb': size_kb,
                        'type': ext,
                        'modified': modified,
                        'marked': False,
                        'category': category,
                        'action': ''
                    })
                    
                    self.status_updated.emit(f"Scanned {len(self.file_list)} files...")
                    
                except (OSError, PermissionError):
                    pass
        
        self.progress_updated.emit(100)
        self.status_updated.emit(f"Found {len(self.file_list)} files - Categorized!")
        self.scan_complete.emit(len(self.file_list))
    
    def cancel(self):
        """Cancel the scan"""
        self.cancel_flag.set()


class FileWhipApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("FileWhip v2.0 - Professional File Organizer")
        self.setGeometry(100, 100, 1400, 900)
        
        # Create menu bar
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("File")
        exit_action = file_menu.addAction("Exit")
        exit_action.triggered.connect(self.close)
        
        # Help menu
        help_menu = menubar.addMenu("Help")
        how_to_use_action = help_menu.addAction("How to Use")
        how_to_use_action.triggered.connect(self.show_how_to_use)
        help_menu.addSeparator()
        about_action = help_menu.addAction("About")
        about_action.triggered.connect(self.show_about)
        
        self.file_list = []
        self.category_summary = {}
        self.move_log = []
        self.known_types = self.get_known_file_types()
        self.program_associations = self.get_program_associations()
        self.unknown_type_cache = {}  # Cache for online lookups
        self.organize_plan = []  # Store organize plan for apply/undo
        self.applied_plan = False  # Track if plan has been applied
        self.operation_mode = "Move"  # "Move" or "Copy"
        self.created_directories = set()  # Track directories created by the tool
        
        # Thread-safe communication
        self.current_thread = None
        self.cancel_flag = threading.Event()
        
        # Configuration
        self.config = self.load_config()
        
        self.setup_ui()
    
    def load_config(self):
        """Load configuration from file or use defaults"""
        config_file = os.path.join(os.path.dirname(__file__), "config.json")
        default_config = {
            "cutoff_date": "2023-01-01",
            "clutter_types": ["tmp", "temp", "cache", "log", "bak"],
            "use_hash_duplication": False
        }
        
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r') as f:
                    return {**default_config, **json.load(f)}
            except:
                return default_config
        return default_config
    
    def save_config(self):
        """Save configuration to file"""
        config_file = os.path.join(os.path.dirname(__file__), "config.json")
        try:
            with open(config_file, 'w') as f:
                json.dump(self.config, f, indent=2)
        except:
            pass
    
    def get_known_file_types(self):
        """Return dictionary of known file types and their categories"""
        return {
            # Music
            'mp3': 'Music', 'wav': 'Music', 'flac': 'Music', 'aac': 'Music', 'm4a': 'Music',
            'ogg': 'Music', 'wma': 'Music', 'aiff': 'Music', 'alac': 'Music', 'opus': 'Music',
            'mid': 'Music', 'midi': 'Music', 'amr': 'Music', '3gp': 'Music', 'ra': 'Music',
            'au': 'Music', 'ac3': 'Music', 'dts': 'Music', 'mka': 'Music', 'ape': 'Music',
            'wv': 'Music', 'tak': 'Music', 'tta': 'Music', 'dsf': 'Music', 'dff': 'Music',
            # Video
            'mp4': 'Video', 'avi': 'Video', 'mkv': 'Video', 'mov': 'Video', 'webm': 'Video',
            'flv': 'Video', 'wmv': 'Video', 'm4v': 'Video', '3gp': 'Video', '3g2': 'Video',
            'rm': 'Video', 'rmvb': 'Video', 'asf': 'Video', 'divx': 'Video', 'xvid': 'Video',
            'ts': 'Video', 'mts': 'Video', 'm2ts': 'Video', 'vob': 'Video', 'ogv': 'Video',
            'drc': 'Video', 'mxf': 'Video', 'roq': 'Video', 'nsv': 'Video', 'f4v': 'Video',
            'f4p': 'Video', 'f4a': 'Video', 'f4b': 'Video',
            # Documents
            'doc': 'Document', 'docx': 'Document', 'pdf': 'Document', 'txt': 'Document',
            'rtf': 'Document', 'odt': 'Document', 'ott': 'Document', 'sxw': 'Document',
            'tex': 'Document', 'wpd': 'Document', 'wps': 'Document', 'abw': 'Document',
            'pages': 'Document', 'epub': 'Document', 'mobi': 'Document', 'azw': 'Document',
            'azw3': 'Document', 'kfx': 'Document', 'lit': 'Document', 'pdb': 'Document',
            'tcr': 'Document', 'lrf': 'Document', 'djvu': 'Document', 'fb2': 'Document',
            # Spreadsheets
            'xls': 'Spreadsheet', 'xlsx': 'Spreadsheet', 'xlsm': 'Spreadsheet', 'xlsb': 'Spreadsheet',
            'csv': 'Spreadsheet', 'ods': 'Spreadsheet', 'ots': 'Spreadsheet', 'fods': 'Spreadsheet',
            'sxc': 'Spreadsheet', 'stc': 'Spreadsheet', 'numbers': 'Spreadsheet',
            # Presentations
            'ppt': 'Presentation', 'pptx': 'Presentation', 'pps': 'Presentation', 'ppsx': 'Presentation',
            'pot': 'Presentation', 'potx': 'Presentation', 'odp': 'Document', 'otp': 'Document',
            'fodp': 'Document', 'sxi': 'Presentation', 'sti': 'Presentation', 'key': 'Presentation',
            # Images
            'jpg': 'Image', 'jpeg': 'Image', 'png': 'Image', 'gif': 'Image', 'bmp': 'Image',
            'tiff': 'Image', 'tif': 'Image', 'webp': 'Image', 'svg': 'Image', 'ico': 'Image',
            'raw': 'Image', 'cr2': 'Image', 'nef': 'Image', 'arw': 'Image', 'dng': 'Image',
            'orf': 'Image', 'rw2': 'Image', 'pef': 'Image', 'raf': 'Image', 'sr2': 'Image',
            'psd': 'Image', 'psb': 'Image', 'ai': 'Image', 'eps': 'Image', 'xcf': 'Image',
            'heic': 'Image', 'heif': 'Image', 'avif': 'Image', 'jxl': 'Image', 'apng': 'Image',
            # Web
            'html': 'Web', 'htm': 'Web', 'xml': 'Web', 'xhtml': 'Web', 'css': 'Web',
            'scss': 'Web', 'sass': 'Web', 'less': 'Web', 'js': 'Web', 'json': 'Web',
            'yaml': 'Web', 'yml': 'Web', 'toml': 'Web', 'md': 'Web', 'markdown': 'Web',
            'rss': 'Web', 'atom': 'Web', 'xsl': 'Web', 'xslt': 'Web', 'xslfo': 'Web',
            # Config/System
            'ini': 'Config', 'cfg': 'Config', 'conf': 'Config', 'json': 'Config',
            'yaml': 'Config', 'yml': 'Config', 'toml': 'Config', 'xml': 'Config',
            'properties': 'Config', 'env': 'Config', 'bat': 'System', 'sh': 'System',
            'bash': 'System', 'zsh': 'System', 'fish': 'System', 'ps1': 'System',
            'vbs': 'System', 'cmd': 'System', 'reg': 'System', 'inf': 'System',
            # Archives
            'zip': 'Archive', 'rar': 'Archive', '7z': 'Archive', 'tar': 'Archive',
            'gz': 'Archive', 'bz2': 'Archive', 'xz': 'Archive', 'lzma': 'Archive',
            'cab': 'Archive', 'iso': 'Archive', 'img': 'Archive', 'dmg': 'Archive',
            'vhd': 'Archive', 'vmdk': 'Archive', 'ova': 'Archive', 'ovf': 'Archive',
            'apk': 'Archive', 'ipa': 'Archive', 'deb': 'Archive', 'rpm': 'Archive',
            'msi': 'Archive', 'jar': 'Archive', 'war': 'Archive', 'ear': 'Archive',
            # Code
            'py': 'Code', 'pyw': 'Code', 'pyc': 'Code', 'pyo': 'Code', 'pyd': 'Code',
            'js': 'Code', 'ts': 'Code', 'jsx': 'Code', 'tsx': 'Code', 'vue': 'Code',
            'java': 'Code', 'class': 'Code', 'jar': 'Code', 'war': 'Code', 'kt': 'Code',
            'kts': 'Code', 'scala': 'Code', 'go': 'Code', 'rs': 'Code', 'c': 'Code',
            'cpp': 'Code', 'cc': 'Code', 'cxx': 'Code', 'h': 'Code', 'hpp': 'Code',
            'hxx': 'Code', 'cs': 'Code', 'vb': 'Code', 'vba': 'Code', 'php': 'Code',
            'rb': 'Code', 'pl': 'Code', 'pm': 'Code', 'lua': 'Code', 'r': 'Code',
            'swift': 'Code', 'm': 'Code', 'mm': 'Code', 'h': 'Code', 'sql': 'Code',
            'sh': 'Code', 'bash': 'Code', 'ps1': 'Code', 'bat': 'Code', 'cmd': 'Code',
            # 3D/CAD
            '3ds': '3D/CAD', 'blend': '3D/CAD', 'fbx': '3D/CAD', 'obj': '3D/CAD',
            'stl': '3D/CAD', 'step': '3D/CAD', 'stp': '3D/CAD', 'iges': '3D/CAD',
            'igs': '3D/CAD', 'dwg': '3D/CAD', 'dxf': '3D/CAD', 'skp': '3D/CAD',
            'max': '3D/CAD', 'ma': '3D/CAD', 'mb': '3D/CAD', 'c4d': '3D/CAD',
            'lwo': '3D/CAD', 'lws': '3D/CAD', 'xsi': '3D/CAD', 'ztl': '3D/CAD',
            'abc': '3D/CAD', 'ply': '3D/CAD', 'pov': '3D/CAD', 'vrml': '3D/CAD',
            'wrl': '3D/CAD', 'dae': '3D/CAD', 'ac': '3D/CAD', 'prt': '3D/CAD',
            'asm': '3D/CAD', 'catpart': '3D/CAD', 'catproduct': '3D/CAD',
            # Fonts
            'ttf': 'Font', 'otf': 'Font', 'woff': 'Font', 'woff2': 'Font', 'eot': 'Font',
            'fon': 'Font', 'fnt': 'Font', 'pfb': 'Font', 'pfm': 'Font', 'afm': 'Font',
            # Credentials
            'pem': 'Credential', 'crt': 'Credential', 'cer': 'Credential', 'key': 'Credential',
            'p12': 'Credential', 'pfx': 'Credential', 'der': 'Credential', 'jks': 'Credential',
            'keystore': 'Credential', 'csr': 'Credential', 'crl': 'Credential',
            # Games
            'sav': 'Game', 'rom': 'Game', 'iso': 'Game', 'bin': 'Game', 'cue': 'Game',
            'ccd': 'Game', 'mds': 'Game', 'mdf': 'Game', 'img': 'Game', 'nrg': 'Game',
            'cdi': 'Game', 'gdi': 'Game', 'ecm': 'Game', 'md5': 'Game', 'sha1': 'Game',
            'sha256': 'Game', 'sfv': 'Game', 'par2': 'Game', 'par': 'Game', 'rev': 'Game',
        }
    
    def setup_ui(self):
        """Setup professional PySide6 UI"""
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # Title
        title_label = QLabel("FileWhip v2.0 - Professional File Organizer")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title_label)
        
        # Control panel
        control_group = QGroupBox("Folder Selection")
        control_layout = QVBoxLayout()
        control_group.setLayout(control_layout)
        
        # Source folder row
        source_layout = QHBoxLayout()
        source_layout.addWidget(QLabel("Source Folder:"))
        self.folder_entry = QLineEdit()
        self.folder_entry.setPlaceholderText("Select folder to scan...")
        source_layout.addWidget(self.folder_entry)
        browse_source_btn = QPushButton("Browse")
        browse_source_btn.clicked.connect(self.browse_folder)
        source_layout.addWidget(browse_source_btn)
        control_layout.addLayout(source_layout)
        
        # Destination folder row
        dest_layout = QHBoxLayout()
        dest_layout.addWidget(QLabel("Destination:"))
        self.dest_entry = QLineEdit()
        self.dest_entry.setPlaceholderText("Select destination folder...")
        dest_layout.addWidget(self.dest_entry)
        browse_dest_btn = QPushButton("Browse")
        browse_dest_btn.clicked.connect(self.browse_destination)
        dest_layout.addWidget(browse_dest_btn)
        control_layout.addLayout(dest_layout)
        
        main_layout.addWidget(control_group)
        
        # Action buttons
        action_group = QGroupBox("Main Actions")
        action_layout = QHBoxLayout()
        action_group.setLayout(action_layout)
        
        self.scan_btn = QPushButton("Scan & Categorize")
        self.scan_btn.setStyleSheet("background-color: #0066cc; color: white; font-weight: bold; padding: 8px;")
        self.scan_btn.clicked.connect(self.scan_and_categorize)
        action_layout.addWidget(self.scan_btn)
        
        self.flag_btn = QPushButton("Flag for Cleanup")
        self.flag_btn.clicked.connect(self.flag_cleanup)
        action_layout.addWidget(self.flag_btn)
        
        self.move_btn = QPushButton("Move Marked Files")
        self.move_btn.setStyleSheet("background-color: #009933; color: white; font-weight: bold; padding: 8px;")
        self.move_btn.clicked.connect(self.move_marked_files)
        action_layout.addWidget(self.move_btn)
        
        self.undo_btn = QPushButton("Undo Moves")
        self.undo_btn.clicked.connect(self.undo_moves)
        action_layout.addWidget(self.undo_btn)
        
        main_layout.addWidget(action_group)
        
        # Additional Tools
        tools_group = QGroupBox("File Management Tools")
        tools_layout = QHBoxLayout()
        tools_group.setLayout(tools_layout)
        
        dup_btn = QPushButton("Find Duplicates")
        dup_btn.clicked.connect(self.find_duplicate_files)
        tools_layout.addWidget(dup_btn)
        
        large_btn = QPushButton("Find Large Files")
        large_btn.clicked.connect(lambda: self.find_large_files())
        tools_layout.addWidget(large_btn)
        
        empty_btn = QPushButton("Find Empty Folders")
        empty_btn.clicked.connect(self.find_empty_folders)
        tools_layout.addWidget(empty_btn)
        
        rename_btn = QPushButton("Batch Rename")
        rename_btn.clicked.connect(self.batch_rename_files)
        tools_layout.addWidget(rename_btn)
        
        music_btn = QPushButton("Organize Music")
        music_btn.clicked.connect(self.organize_music_by_metadata)
        tools_layout.addWidget(music_btn)
        
        export_btn = QPushButton("Export All Tabs")
        export_btn.setStyleSheet("background-color: #9900cc; color: white; font-weight: bold; padding: 8px;")
        export_btn.clicked.connect(self.export_all_tabs)
        tools_layout.addWidget(export_btn)
        
        auto_organize_btn = QPushButton("Auto Organize")
        auto_organize_btn.setStyleSheet("background-color: #ff6600; color: white; font-weight: bold; padding: 8px;")
        auto_organize_btn.clicked.connect(self.auto_organize)
        tools_layout.addWidget(auto_organize_btn)
        
        main_layout.addWidget(tools_group)
        
        # Progress
        self.progress_label = QLabel("Ready")
        main_layout.addWidget(self.progress_label)
        
        self.progress_bar = QProgressBar()
        main_layout.addWidget(self.progress_bar)
        
        # Tab widget
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        # File List Tab
        self.file_list_tab = QWidget()
        file_list_layout = QVBoxLayout(self.file_list_tab)
        
        # Filter bar
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Search:"))
        self.search_entry = QLineEdit()
        self.search_entry.textChanged.connect(self.apply_filter)
        filter_layout.addWidget(self.search_entry)
        
        filter_layout.addWidget(QLabel("Category:"))
        self.category_filter = QComboBox()
        self.category_filter.addItem("All")
        self.category_filter.currentTextChanged.connect(self.apply_filter)
        filter_layout.addWidget(self.category_filter)
        
        self.marked_only_cb = QCheckBox("Marked Only")
        self.marked_only_cb.stateChanged.connect(self.apply_filter)
        filter_layout.addWidget(self.marked_only_cb)
        
        file_list_layout.addLayout(filter_layout)
        
        # Tree widget
        self.tree = QTreeWidget()
        self.tree.setColumnCount(8)
        self.tree.setHeaderLabels(["Filename", "Path", "Size (KB)", "Type", "Modified", "Category", "Programs", "Action"])
        self.tree.setAlternatingRowColors(True)
        self.tree.setSortingEnabled(True)
        self.tree.setItemsExpandable(False)
        file_list_layout.addWidget(self.tree)
        
        self.tab_widget.addTab(self.file_list_tab, "File List")
        
        # Category Summary Tab
        self.summary_tab = QWidget()
        summary_layout = QVBoxLayout(self.summary_tab)
        
        self.summary_tree = QTreeWidget()
        self.summary_tree.setColumnCount(3)
        self.summary_tree.setHeaderLabels(["Category", "File Count", "Total Size (KB)"])
        self.summary_tree.setAlternatingRowColors(True)
        self.summary_tree.setSortingEnabled(True)
        self.summary_tree.setItemsExpandable(False)
        summary_layout.addWidget(self.summary_tree)
        
        self.tab_widget.addTab(self.summary_tab, "Category Summary")
        
        # Move Log Tab
        self.log_tab = QWidget()
        log_layout = QVBoxLayout(self.log_tab)
        
        self.log_tree = QTreeWidget()
        self.log_tree.setColumnCount(6)
        self.log_tree.setHeaderLabels(["Filename", "Original Path", "New Path", "Type", "Timestamp", "Status"])
        self.log_tree.setAlternatingRowColors(True)
        self.log_tree.setSortingEnabled(True)
        self.log_tree.setItemsExpandable(False)
        log_layout.addWidget(self.log_tree)
        
        self.tab_widget.addTab(self.log_tab, "Move Log")
        
        # Tool Results Tab
        self.results_tab = QWidget()
        results_layout = QVBoxLayout(self.results_tab)
        
        self.results_tree = QTreeWidget()
        self.results_tree.setColumnCount(4)
        self.results_tree.setHeaderLabels(["Tool", "Item", "Details", "Timestamp"])
        self.results_tree.setAlternatingRowColors(True)
        self.results_tree.setSortingEnabled(True)
        self.results_tree.setItemsExpandable(False)
        results_layout.addWidget(self.results_tree)
        
        clear_results_btn = QPushButton("Clear Results")
        clear_results_btn.clicked.connect(self.clear_tool_results)
        results_layout.addWidget(clear_results_btn)
        
        self.tab_widget.addTab(self.results_tab, "Tool Results")
        
        # Auto Organize Plan Tab
        self.plan_tab = QWidget()
        plan_layout = QVBoxLayout(self.plan_tab)
        
        self.plan_tree = QTreeWidget()
        self.plan_tree.setColumnCount(5)
        self.plan_tree.setHeaderLabels(["Action", "Filename", "Source", "Destination", "Reason"])
        self.plan_tree.setAlternatingRowColors(True)
        self.plan_tree.setSortingEnabled(True)
        self.plan_tree.setItemsExpandable(False)
        plan_layout.addWidget(self.plan_tree)
        
        # Move/Copy selection
        operation_layout = QHBoxLayout()
        operation_layout.addWidget(QLabel("Operation:"))
        self.move_rb = QRadioButton("Move")
        self.move_rb.setChecked(True)
        self.move_rb.toggled.connect(self.on_operation_changed)
        operation_layout.addWidget(self.move_rb)
        
        self.copy_rb = QRadioButton("Copy")
        self.copy_rb.toggled.connect(self.on_operation_changed)
        operation_layout.addWidget(self.copy_rb)
        plan_layout.addLayout(operation_layout)
        
        plan_button_layout = QHBoxLayout()
        self.apply_plan_btn = QPushButton("Apply Plan")
        self.apply_plan_btn.setStyleSheet("background-color: #009933; color: white; font-weight: bold; padding: 8px;")
        self.apply_plan_btn.clicked.connect(self.apply_organize_plan)
        self.apply_plan_btn.setEnabled(False)
        plan_button_layout.addWidget(self.apply_plan_btn)
        
        self.undo_plan_btn = QPushButton("Undo Plan")
        self.undo_plan_btn.setStyleSheet("background-color: #cc0000; color: white; font-weight: bold; padding: 8px;")
        self.undo_plan_btn.clicked.connect(self.undo_organize_plan)
        self.undo_plan_btn.setEnabled(False)
        plan_button_layout.addWidget(self.undo_plan_btn)
        
        plan_layout.addLayout(plan_button_layout)
        
        self.tab_widget.addTab(self.plan_tab, "Auto Organize Plan")
        
        # Auto Organize Log Tab
        self.organize_log_tab = QWidget()
        log_layout = QVBoxLayout(self.organize_log_tab)
        
        self.organize_log_tree = QTreeWidget()
        self.organize_log_tree.setColumnCount(4)
        self.organize_log_tree.setHeaderLabels(["Status", "Filename", "Source", "Destination"])
        self.organize_log_tree.setAlternatingRowColors(True)
        self.organize_log_tree.setSortingEnabled(True)
        self.organize_log_tree.setItemsExpandable(False)
        log_layout.addWidget(self.organize_log_tree)
        
        clear_log_btn = QPushButton("Clear Log")
        clear_log_btn.clicked.connect(lambda: self.organize_log_tree.clear())
        log_layout.addWidget(clear_log_btn)
        
        self.tab_widget.addTab(self.organize_log_tab, "Auto Organize Log")
        
        # Status bar
        self.status_label = QLabel("Ready")
        self.statusBar().addWidget(self.status_label)
    
    def clear_tool_results(self):
        """Clear the tool results tree"""
        self.results_tree.clear()
    
    def on_operation_changed(self):
        """Handle operation mode change (Move/Copy)"""
        if self.move_rb.isChecked():
            self.operation_mode = "Move"
        elif self.copy_rb.isChecked():
            self.operation_mode = "Copy"
    
    def show_how_to_use(self):
        """Display How to Use documentation"""
        help_text = """
        <h2>FileWhip v2.0 - Professional File Organizer</h2>
        
        <h3>Getting Started</h3>
        <ol>
            <li><b>Scan a Folder:</b> Click "Browse" to select a source folder to scan. FileWhip will analyze all files in the folder and subfolders.</li>
            <li><b>Select Destination:</b> Click "Browse Destination" to select where you want files to be organized.</li>
            <li><b>Review Files:</b> View the scanned files in the main table, organized by category and file type.</li>
        </ol>
        
        <h3>File Management Tools</h3>
        <ul>
            <li><b>Find Duplicates:</b> Locate duplicate files using multiple criteria (exact hash, filename, size, date, similar names). Shows all locations for each duplicate group.</li>
            <li><b>Find Large Files:</b> Identify files larger than a specified size (default: 100 MB).</li>
            <li><b>Find Empty Folders:</b> Discover empty folders in the scanned directory.</li>
            <li><b>Batch Rename:</b> Rename multiple files using find and replace patterns.</li>
            <li><b>Organize Music:</b> Organize music files by artist/album using metadata (requires mutagen library).</li>
        </ul>
        
        <h3>Auto Organize</h3>
        <p>The Auto Organize feature creates an intelligent plan to organize your files:</p>
        <ul>
            <li><b>Duplicates:</b> Moves duplicate files to a Duplicates folder in the destination.</li>
            <li><b>Large Files:</b> Moves files >100 MB to a Large_Files folder.</li>
            <li><b>Music Files:</b> Organizes by Artist/Album using metadata.</li>
            <li><b>Category-based:</b> Organizes files by type (Documents, Images, Videos, etc.) considering the source folder name.</li>
            <li><b>Unknown Files:</b> Moves unknown file types to Unknown_Files for review.</li>
        </ul>
        <p><b>Operation Modes:</b></p>
        <ul>
            <li><b>Move:</b> Moves files from source to destination (cleaning up source).</li>
            <li><b>Copy:</b> Copies files to destination (preserving source).</li>
        </ul>
        <p><b>Apply & Undo:</b> Review the plan in the Auto Organize Plan tab, then apply changes. Use Undo to revert (time travel - removes files/directories created by the tool).</p>
        
        <h3>Tabs</h3>
        <ul>
            <li><b>Files:</b> Main view of scanned files with category filtering.</li>
            <li><b>Summary:</b> Overview of file statistics by category.</li>
            <li><b>Tool Results:</b> Results from file management tools in table format.</li>
            <li><b>Auto Organize Plan:</b> Review the proposed organization plan before applying.</li>
            <li><b>Auto Organize Log:</b> Log of changes and failures after applying the plan.</li>
        </ul>
        
        <h3>Export</h3>
        <p>Click "Export All Tabs" to export all tabs to XLSX, CSV, or Google Sheets format.</p>
        
        <h3>Important Notes</h3>
        <ul>
            <li>The tool only moves files (never deletes) except when using Undo to revert changes.</li>
            <li>Duplicate detection shows ALL locations where duplicate files exist.</li>
            <li>Auto Organize considers the source folder name when determining destination.</li>
            <li>Undo works as "time travel" - restores files and removes empty directories created by the tool.</li>
        </ul>
        """
        
        dialog = QDialog(self)
        dialog.setWindowTitle("How to Use - FileWhip")
        dialog.resize(800, 600)
        
        layout = QVBoxLayout()
        
        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setHtml(help_text)
        layout.addWidget(text_edit)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def show_about(self):
        """Display About dialog with tool metadata"""
        dialog = QDialog(self)
        dialog.setWindowTitle("About - FileWhip")
        dialog.setMinimumSize(600, 400)
        
        layout = QVBoxLayout(dialog)
        
        label = QLabel(dialog)
        label.setText("FileWhip - Whip your files into shape\nVersion 2.0\n\nCreator: Nick Garofalo\nCopyright: 2026 Nick Garofalo. All rights reserved.\n\nDescription:\nFileWhip is a professional file organization tool designed to help you manage and organize your files efficiently. With intelligent auto-organization, duplicate detection, and comprehensive file management tools, FileWhip makes file management a breeze.\n\nKey Features:\n- Intelligent Auto Organize with Move/Copy options\n- Duplicate detection with multiple criteria (hash, filename, size, date, similar names)\n- Large files identification (>100 MB by default)\n- Empty folder discovery and cleanup\n- Batch rename with find and replace patterns\n- Music organization by artist/album using metadata\n- Time Travel Undo - revert all changes and remove empty directories\n- Export to XLSX, CSV, or Google Sheets format\n- Category-based file organization with source folder consideration\n- Tool Results tab with detailed logs\n- Auto Organize Log for tracking changes and failures\n\nTechnology:\n- Built with Python 3\n- GUI Framework: PySide6 (Qt6)\n- File Operations: shutil, os\n- Hash Calculation: hashlib (MD5)\n- Excel Export: openpyxl\n- Audio Metadata: mutagen (optional)\n- HTTP Requests: requests\n\nSystem Requirements:\n- Windows 10/11 or compatible Linux distribution\n- Python 3.8 or higher\n- Minimum 4GB RAM recommended\n\nDependencies:\n- PySide6\n- openpyxl\n- mutagen (optional, for music metadata)\n- requests\n\nLicense:\nThis software is provided as-is for personal and professional use.\n\nFileWhip - Whip your files into shape")
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignTop | Qt.AlignLeft)
        layout.addWidget(label)
        
        close_btn = QPushButton("Close", dialog)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def export_all_tabs(self):
        """Export all tabs to selected format"""
        # Create export format selection dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Export Format Selection")
        dialog.resize(400, 300)
        layout = QVBoxLayout()
        
        layout.addWidget(QLabel("Select export format:"))
        
        xlsx_rb = QRadioButton("Excel (.xlsx) - Single file with multiple sheets")
        xlsx_rb.setChecked(True)
        layout.addWidget(xlsx_rb)
        
        csv_rb = QRadioButton("Multiple CSV files - One file per tab")
        layout.addWidget(csv_rb)
        
        gsheets_rb = QRadioButton("Google Sheets - Requires API credentials")
        layout.addWidget(gsheets_rb)
        
        def execute_export():
            if xlsx_rb.isChecked():
                self.export_to_xlsx()
            elif csv_rb.isChecked():
                self.export_to_csv()
            elif gsheets_rb.isChecked():
                self.export_to_google_sheets()
            dialog.accept()
        
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(execute_export)
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def export_to_xlsx(self):
        """Export all tabs to a single Excel file with multiple sheets"""
        file_path, _ = QFileDialog.getSaveFileName(self, "Export to Excel", "", "Excel files (*.xlsx);;All files (*.*)")
        if not file_path:
            return
        
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Export File List tab
            if self.file_list:
                ws = wb.create_sheet("File List")
                ws.append(["Filename", "Path", "Size (KB)", "Type", "Modified", "Category", "Programs", "Action"])
                for file_info in self.file_list:
                    ext = file_info['type']
                    programs = self.program_associations.get(ext, 'Unknown')
                    if programs == 'Unknown' and ext:
                        programs = self.lookup_file_type_online(ext)
                    ws.append([
                        file_info['filename'],
                        file_info['path'],
                        file_info['size_kb'],
                        ext,
                        file_info['modified'],
                        file_info['category'],
                        programs,
                        file_info.get('action', '')
                    ])
            
            # Export Category Summary tab
            if self.category_summary:
                ws = wb.create_sheet("Category Summary")
                ws.append(["Category", "File Count", "Total Size (KB)"])
                for category, data in sorted(self.category_summary.items()):
                    ws.append([category, data['count'], data['size']])
            
            # Export Move Log tab
            if self.move_log:
                ws = wb.create_sheet("Move Log")
                ws.append(["Filename", "Original Path", "New Path", "Type", "Timestamp", "Status"])
                for entry in self.move_log:
                    ws.append([
                        entry.get('filename', ''),
                        entry.get('original_path', ''),
                        entry.get('new_path', ''),
                        entry.get('type', ''),
                        entry.get('timestamp', ''),
                        entry.get('status', '')
                    ])
            
            # Export Tool Results tab
            if self.results_tree.topLevelItemCount() > 0:
                ws = wb.create_sheet("Tool Results")
                ws.append(["Tool", "Item", "Details", "Timestamp"])
                for i in range(self.results_tree.topLevelItemCount()):
                    item = self.results_tree.topLevelItem(i)
                    ws.append([
                        item.text(0),
                        item.text(1),
                        item.text(2),
                        item.text(3)
                    ])
            
            # Export Auto Organize Plan tab
            if self.plan_tree.topLevelItemCount() > 0:
                ws = wb.create_sheet("Auto Organize Plan")
                ws.append(["Action", "Filename", "Source", "Destination", "Reason"])
                for i in range(self.plan_tree.topLevelItemCount()):
                    item = self.plan_tree.topLevelItem(i)
                    ws.append([
                        item.text(0),
                        item.text(1),
                        item.text(2),
                        item.text(3),
                        item.text(4)
                    ])
            
            wb.save(file_path)
            QMessageBox.information(self, "Export Complete", f"Successfully exported to {file_path}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {e}")
    
    def export_to_csv(self):
        """Export all tabs to multiple CSV files"""
        folder = QFileDialog.getExistingDirectory(self, "Select Export Folder")
        if not folder:
            return
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Export File List
            if self.file_list:
                file_path = os.path.join(folder, f"FileList_{timestamp}.csv")
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Filename", "Path", "Size (KB)", "Type", "Modified", "Category", "Programs", "Action"])
                    for file_info in self.file_list:
                        ext = file_info['type']
                        programs = self.program_associations.get(ext, 'Unknown')
                        if programs == 'Unknown' and ext:
                            programs = self.lookup_file_type_online(ext)
                        writer.writerow([
                            file_info['filename'],
                            file_info['path'],
                            file_info['size_kb'],
                            ext,
                            file_info['modified'],
                            file_info['category'],
                            programs,
                            file_info.get('action', '')
                        ])
            
            # Export Category Summary
            if self.category_summary:
                file_path = os.path.join(folder, f"CategorySummary_{timestamp}.csv")
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Category", "File Count", "Total Size (KB)"])
                    for category, data in sorted(self.category_summary.items()):
                        writer.writerow([category, data['count'], data['size']])
            
            # Export Move Log
            if self.move_log:
                file_path = os.path.join(folder, f"MoveLog_{timestamp}.csv")
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Filename", "Original Path", "New Path", "Type", "Timestamp", "Status"])
                    for entry in self.move_log:
                        writer.writerow([
                            entry.get('filename', ''),
                            entry.get('original_path', ''),
                            entry.get('new_path', ''),
                            entry.get('type', ''),
                            entry.get('timestamp', ''),
                            entry.get('status', '')
                        ])
            
            # Export Tool Results
            if self.results_tree.topLevelItemCount() > 0:
                file_path = os.path.join(folder, f"ToolResults_{timestamp}.csv")
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Tool", "Item", "Details", "Timestamp"])
                    for i in range(self.results_tree.topLevelItemCount()):
                        item = self.results_tree.topLevelItem(i)
                        writer.writerow([
                            item.text(0),
                            item.text(1),
                            item.text(2),
                            item.text(3)
                        ])
            
            # Export Auto Organize Plan
            if self.plan_tree.topLevelItemCount() > 0:
                file_path = os.path.join(folder, f"AutoOrganizePlan_{timestamp}.csv")
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Action", "Filename", "Source", "Destination", "Reason"])
                    for i in range(self.plan_tree.topLevelItemCount()):
                        item = self.plan_tree.topLevelItem(i)
                        writer.writerow([
                            item.text(0),
                            item.text(1),
                            item.text(2),
                            item.text(3),
                            item.text(4)
                        ])
            
            QMessageBox.information(self, "Export Complete", f"Successfully exported CSV files to {folder}")
        except Exception as e:
            QMessageBox.warning(self, "Export Error", f"Failed to export: {e}")
    
    def auto_organize(self):
        """Analyze files and create an auto-organization plan"""
        if not self.file_list:
            QMessageBox.warning(self, "Warning", "No files to analyze. Scan a folder first.")
            return
        
        dest_folder = self.dest_entry.text()
        if not dest_folder:
            QMessageBox.warning(self, "Warning", "Please select a destination folder first.")
            return
        
        self.status_label.setText("Analyzing files for auto-organization...")
        self.plan_tree.clear()
        self.organize_plan = []
        self.applied_plan = False
        
        cutoff_date = self.config.get("cutoff_date", "2023-01-01")
        cutoff_datetime = datetime.strptime(cutoff_date, "%Y-%m-%d")
        
        # Pre-calculate all file hashes for duplicate detection
        hash_dict = defaultdict(list)
        for file_info in self.file_list:
            file_hash = self.calculate_file_hash(file_info['path'])
            if file_hash:
                hash_dict[file_hash].append(file_info)
        
        # Mark duplicates
        duplicate_hashes = {h for h, files in hash_dict.items() if len(files) > 1}
        
        # Analyze files and create plan
        for file_info in self.file_list:
            action = None
            destination = None
            reason = None
            
            # Check for duplicates (by hash)
            file_hash = self.calculate_file_hash(file_info['path'])
            if file_hash and file_hash in duplicate_hashes:
                # This is a duplicate, move to Duplicates folder in destination
                duplicates_folder = os.path.join(dest_folder, "Duplicates")
                action = "Move"
                destination = duplicates_folder
                reason = "Duplicate file detected"
                self.organize_plan.append({
                    'action': action,
                    'filename': file_info['filename'],
                    'source': file_info['path'],
                    'destination': os.path.join(destination, file_info['filename']),
                    'reason': reason
                })
                continue
            
            # Check for large files (> 100 MB)
            if file_info['size_kb'] > 100 * 1024:
                action = "Move"
                destination = os.path.join(dest_folder, "Large_Files")
                reason = f"Large file ({file_info['size_kb']/1024:.1f} MB)"
                self.organize_plan.append({
                    'action': action,
                    'filename': file_info['filename'],
                    'source': file_info['path'],
                    'destination': os.path.join(destination, file_info['filename']),
                    'reason': reason
                })
                continue
            
            # Check for music files with metadata
            if file_info['category'] == 'Music' and MUTAGEN_AVAILABLE:
                metadata = self.extract_audio_metadata(file_info['path'])
                if metadata and metadata['artist'] != 'Unknown' and metadata['album'] != 'Unknown':
                    artist = metadata['artist'].replace('/', '_').replace('\\', '_')
                    album = metadata['album'].replace('/', '_').replace('\\', '_')
                    action = "Move"
                    destination = os.path.join(dest_folder, "Music", artist, album)
                    reason = f"Music: {artist} - {album}"
                    self.organize_plan.append({
                        'action': action,
                        'filename': file_info['filename'],
                        'source': file_info['path'],
                        'destination': os.path.join(destination, file_info['filename']),
                        'reason': reason
                    })
                    continue
            
            # Organize by category (consider both file type and source folder)
            source_folder_name = os.path.basename(os.path.dirname(file_info['path']))
            
            if file_info['category'] != 'Unknown':
                # Use file extension category, but also consider source folder
                action = "Move"
                # If source folder has a meaningful name, use it as subcategory
                if source_folder_name and source_folder_name not in ['.', 'source', 'files']:
                    destination = os.path.join(dest_folder, file_info['category'].replace(' ', '_'), source_folder_name)
                    reason = f"Category: {file_info['category']} (from {source_folder_name})"
                else:
                    destination = os.path.join(dest_folder, file_info['category'].replace(' ', '_'))
                    reason = f"Category: {file_info['category']}"
                self.organize_plan.append({
                    'action': action,
                    'filename': file_info['filename'],
                    'source': file_info['path'],
                    'destination': os.path.join(destination, file_info['filename']),
                    'reason': reason
                })
            else:
                # Unknown files go to review folder, but organize by source folder
                action = "Move"
                if source_folder_name and source_folder_name not in ['.', 'source', 'files']:
                    destination = os.path.join(dest_folder, "Unknown_Files", source_folder_name)
                    reason = f"Unknown file type (from {source_folder_name})"
                else:
                    destination = os.path.join(dest_folder, "Unknown_Files")
                    reason = "Unknown file type"
                self.organize_plan.append({
                    'action': action,
                    'filename': file_info['filename'],
                    'source': file_info['path'],
                    'destination': os.path.join(destination, file_info['filename']),
                    'reason': reason
                })
        
        # Populate plan tree
        for plan_item in self.organize_plan:
            item = QTreeWidgetItem(self.plan_tree)
            item.setText(0, plan_item['action'])
            item.setText(1, plan_item['filename'])
            item.setText(2, plan_item['source'])
            item.setText(3, plan_item['destination'])
            item.setText(4, plan_item['reason'])
        
        # Enable buttons
        if self.organize_plan:
            self.apply_plan_btn.setEnabled(True)
            self.status_label.setText(f"Plan created: {len(self.organize_plan)} actions to perform")
        else:
            self.status_label.setText("No actions suggested")
        
        self.tab_widget.setCurrentWidget(self.plan_tab)
    
    def apply_organize_plan(self):
        """Apply the auto-organization plan"""
        if not self.organize_plan:
            QMessageBox.warning(self, "Warning", "No plan to apply.")
            return
        
        self.status_label.setText("Applying organization plan...")
        self.progress_bar.setValue(0)
        self.organize_log_tree.clear()  # Clear log before applying
        self.created_directories = set()  # Reset created directories
        
        applied_count = 0
        failed_count = 0
        self.applied_plan_moves = []  # Store for undo
        
        total_actions = len(self.organize_plan)
        operation = self.operation_mode.lower()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for i, plan_item in enumerate(self.organize_plan):
            source = plan_item['source']
            destination = plan_item['destination']
            filename = plan_item['filename']
            
            try:
                # Create destination directory if it doesn't exist
                dest_dir = os.path.dirname(destination)
                if not os.path.exists(dest_dir):
                    os.makedirs(dest_dir, exist_ok=True)
                    self.created_directories.add(dest_dir)
                    # Also add parent directories that were created
                    current_dir = dest_dir
                    while current_dir != os.path.dirname(current_dir):
                        current_dir = os.path.dirname(current_dir)
                        if not os.path.exists(current_dir):
                            self.created_directories.add(current_dir)
                
                if operation == "move":
                    # Move file
                    if os.path.exists(source) and not os.path.exists(destination):
                        shutil.move(source, destination)
                        self.applied_plan_moves.append({
                            'source': source,
                            'destination': destination,
                            'operation': 'move'
                        })
                        applied_count += 1
                        # Add to log
                        item = QTreeWidgetItem(self.organize_log_tree)
                        item.setText(0, "Success")
                        item.setText(1, filename)
                        item.setText(2, source)
                        item.setText(3, destination)
                    elif os.path.exists(destination):
                        # File already exists at destination
                        failed_count += 1
                        # Add to log
                        item = QTreeWidgetItem(self.organize_log_tree)
                        item.setText(0, "Failed")
                        item.setText(1, filename)
                        item.setText(2, source)
                        item.setText(3, "File already exists at destination")
                elif operation == "copy":
                    # Copy file
                    if os.path.exists(source):
                        if not os.path.exists(destination):
                            shutil.copy2(source, destination)
                            self.applied_plan_moves.append({
                                'source': source,
                                'destination': destination,
                                'operation': 'copy'
                            })
                            applied_count += 1
                            # Add to log
                            item = QTreeWidgetItem(self.organize_log_tree)
                            item.setText(0, "Success")
                            item.setText(1, filename)
                            item.setText(2, source)
                            item.setText(3, destination)
                        else:
                            # File already exists at destination
                            failed_count += 1
                            # Add to log
                            item = QTreeWidgetItem(self.organize_log_tree)
                            item.setText(0, "Failed")
                            item.setText(1, filename)
                            item.setText(2, source)
                            item.setText(3, "File already exists at destination")
            except Exception as e:
                failed_count += 1
                print(f"Error {operation}ing {source}: {e}")
                # Add to log
                item = QTreeWidgetItem(self.organize_log_tree)
                item.setText(0, "Failed")
                item.setText(1, filename)
                item.setText(2, source)
                item.setText(3, f"Error: {str(e)}")
            
            # Update progress
            progress = int(((i + 1) / total_actions) * 100)
            self.progress_bar.setValue(progress)
            self.status_label.setText(f"Applying plan: {applied_count}/{total_actions} files {operation}ed...")
        
        self.applied_plan = True
        self.apply_plan_btn.setEnabled(False)
        self.undo_plan_btn.setEnabled(True)
        
        self.progress_bar.setValue(100)
        self.status_label.setText(f"Plan applied: {applied_count} files {operation}ed, {failed_count} failed")
        
        # Switch to log tab
        self.tab_widget.setCurrentWidget(self.organize_log_tab)
        
        QMessageBox.information(self, "Plan Applied", 
            f"Organization plan applied successfully!\n\n"
            f"Operation: {operation.capitalize()}\n"
            f"Files {operation}ed: {applied_count}\n"
            f"Failed: {failed_count}")
    
    def undo_organize_plan(self):
        """Undo the applied organization plan"""
        if not self.applied_plan:
            QMessageBox.warning(self, "Warning", "No plan to undo or plan not yet applied.")
            return
        
        if not hasattr(self, 'applied_plan_moves') or not self.applied_plan_moves:
            QMessageBox.warning(self, "Warning", "No moves to undo.")
            return
        
        # Check what operation was used
        operation = self.applied_plan_moves[0].get('operation', 'move') if self.applied_plan_moves else 'move'
        
        if operation == 'copy':
            reply = QMessageBox.question(self, "Confirm Undo", 
                "Are you sure you want to undo the copy operation? This will DELETE all copied files (original files remain in place).",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        else:
            reply = QMessageBox.question(self, "Confirm Undo", 
                "Are you sure you want to undo all file moves? This will move files back to their original locations.",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.No:
            return
        
        self.status_label.setText("Undoing organization plan...")
        self.progress_bar.setValue(0)
        
        undone_count = 0
        failed_count = 0
        
        total_moves = len(self.applied_plan_moves)
        
        for i, move in enumerate(self.applied_plan_moves):
            if operation == 'copy':
                # Undo copy: delete the copied files
                file_to_delete = move['destination']
                try:
                    if os.path.exists(file_to_delete):
                        os.remove(file_to_delete)
                        undone_count += 1
                except Exception as e:
                    failed_count += 1
                    print(f"Error deleting {file_to_delete}: {e}")
            else:
                # Undo move: move files back to original location
                source = move['destination']  # Current location (was destination)
                destination = move['source']  # Original location (was source)
                
                try:
                    # Move back to original location
                    if os.path.exists(source) and not os.path.exists(destination):
                        dest_dir = os.path.dirname(destination)
                        os.makedirs(dest_dir, exist_ok=True)
                        shutil.move(source, destination)
                        undone_count += 1
                    elif os.path.exists(destination):
                        # File already exists at original location
                        failed_count += 1
                except Exception as e:
                    failed_count += 1
                    print(f"Error undoing move {source}: {e}")
            
            # Update progress
            progress = int(((i + 1) / total_moves) * 100)
            self.progress_bar.setValue(progress)
            self.status_label.setText(f"Undoing: {undone_count}/{total_moves} files restored...")
        
        self.applied_plan = False
        self.applied_plan_moves = []
        self.apply_plan_btn.setEnabled(True)
        self.undo_plan_btn.setEnabled(False)
        
        # Remove empty directories created by the tool (time travel cleanup)
        directories_removed = 0
        # Sort directories in reverse order (deepest first) to ensure we can remove them
        for directory in sorted(self.created_directories, key=lambda x: x.count(os.sep), reverse=True):
            try:
                if os.path.exists(directory) and not os.listdir(directory):
                    os.rmdir(directory)
                    directories_removed += 1
            except Exception as e:
                print(f"Could not remove directory {directory}: {e}")
        
        self.created_directories = set()  # Clear tracked directories
        
        self.progress_bar.setValue(100)
        
        if operation == 'copy':
            self.status_label.setText(f"Plan undone: {undone_count} copied files deleted, {directories_removed} directories removed, {failed_count} failed")
            QMessageBox.information(self, "Plan Undone", 
                f"Copy operation undone successfully!\n\n"
                f"Files deleted: {undone_count}\n"
                f"Directories removed: {directories_removed}\n"
                f"Failed: {failed_count}")
        else:
            self.status_label.setText(f"Plan undone: {undone_count} files restored, {directories_removed} directories removed, {failed_count} failed")
            QMessageBox.information(self, "Plan Undone", 
                f"Organization plan undone successfully!\n\n"
                f"Files restored: {undone_count}\n"
                f"Directories removed: {directories_removed}\n"
                f"Failed: {failed_count}")
    
    def export_to_google_sheets(self):
        """Export to Google Sheets (requires API credentials)"""
        QMessageBox.information(self, "Google Sheets Export", 
            "Google Sheets export requires API credentials.\n\n"
            "To enable this feature:\n"
            "1. Create a Google Cloud project\n"
            "2. Enable Google Sheets API\n"
            "3. Create service account credentials\n"
            "4. Install gspread library: pip install gspread\n"
            "5. Place credentials.json in the application directory\n\n"
            "For now, please use Excel or CSV export.")
    
    def new_scan(self):
        """Start a new scan"""
        self.file_list = []
        self.category_summary = {}
        self.tree.clear()
        self.summary_tree.clear()
        self.clear_tool_results()
        self.plan_tree.clear()
        self.organize_log_tree.clear()
        self.organize_plan = []
        self.applied_plan = False
        self.apply_plan_btn.setEnabled(False)
        self.undo_plan_btn.setEnabled(False)
        self.folder_entry.clear()
        self.dest_entry.clear()
        self.progress_label.setText("Ready")
        self.status_label.setText("Ready")
        self.progress_bar.setValue(0)
    
    def load_scan(self):
        """Load scan results from JSON"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Load Scan", "", "JSON files (*.json);;All files (*.*)")
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                scan_data = json.load(f)
            
            self.file_list = scan_data['files']
            self.folder_entry.setText(scan_data['folder'])
            
            self.refresh_file_list()
            self.refresh_summary()
            
            QMessageBox.information(self, "Success", f"Loaded {len(self.file_list)} files from your saved scan")
            self.status_label.setText("Scan loaded")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Couldn't load: {e}")
    
    def save_scan(self):
        """Save scan results to JSON"""
        if not self.file_list:
            QMessageBox.warning(self, "Warning", "No scan results to save yet.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Scan", "", "JSON files (*.json);;All files (*.*)")
        
        if not file_path:
            return
        
        try:
            scan_data = {
                'timestamp': datetime.now().isoformat(),
                'folder': self.folder_entry.text(),
                'files': self.file_list
            }
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(scan_data, f, indent=2)
            
            QMessageBox.information(self, "Success", "Your scan results have been saved")
            self.status_label.setText("Scan saved")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Couldn't save: {e}")
    
    def export_to_csv(self):
        """Export file list to CSV"""
        if not self.file_list:
            QMessageBox.warning(self, "Warning", "No files to export yet. Please scan a folder first.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(self, "Export to CSV", "", "CSV files (*.csv);;All files (*.*)")
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Filename", "Path", "Size (KB)", "Type", "Modified", "Category"])
                
                for file_info in self.file_list:
                    writer.writerow([
                        file_info['filename'],
                        file_info['path'],
                        file_info['size_kb'],
                        file_info['type'],
                        file_info['modified'],
                        file_info['category']
                    ])
            
            QMessageBox.information(self, "Success", "Your file list has been exported successfully")
            self.status_label.setText("Export completed")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Couldn't export: {e}")
    
    def categorize_files(self):
        """Categorize files by type"""
        if not self.file_list:
            QMessageBox.warning(self, "Warning", "No files found yet. Please scan a folder first.")
            return
        
        self.progress_label.setText("Categorizing...")
        self.status_label.setText("Sorting files by type...")
        
        known_types = self.get_known_file_types()
        
        for file in self.file_list:
            ext = file['type']
            if ext in known_types:
                file['category'] = known_types[ext]
            else:
                file['category'] = 'Unknown'
        
        self.progress_label.setText("Done!")
        self.status_label.setText("Files sorted successfully")
        self.refresh_file_list()
        self.refresh_summary()
    
    def move_marked_files(self):
        """Move marked files to destination folder"""
        if not self.file_list:
            QMessageBox.warning(self, "Warning", "No files to move yet. Please scan a folder first.")
            return
        
        dest_folder = self.dest_entry.text()
        if not dest_folder or not os.path.exists(dest_folder):
            QMessageBox.warning(self, "Error", "Please select a valid destination folder")
            return
        
        marked_files = [f for f in self.file_list if f.get('marked')]
        if not marked_files:
            QMessageBox.warning(self, "Warning", "No files are marked for moving. Use 'Flag for Cleanup' first.")
            return
        
        reply = QMessageBox.question(self, "Confirm", f"Move {len(marked_files)} files to destination?", 
                                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No:
            return
        
        self.progress_label.setText("Moving...")
        self.status_label.setText("Moving files...")
        
        moved_count = 0
        log_entries = []
        
        for file_info in marked_files:
            if self.cancel_flag.is_set():
                break
            
            try:
                src = file_info['path']
                filename = file_info['filename']
                dest = os.path.join(dest_folder, filename)
                
                # Handle duplicates
                if os.path.exists(dest):
                    base, ext = os.path.splitext(filename)
                    counter = 1
                    while os.path.exists(os.path.join(dest_folder, f"{base}_{counter}{ext}")):
                        counter += 1
                    dest = os.path.join(dest_folder, f"{base}_{counter}{ext}")
                
                shutil.move(src, dest)
                moved_count += 1
                
                log_entries.append({
                    'filename': filename,
                    'original_path': src,
                    'new_path': dest,
                    'type': file_info['type'],
                    'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'status': 'Moved'
                })
                
                self.status_label.setText(f"Moved {moved_count} files...")
                
            except Exception as e:
                log_entries.append({
                    'filename': file_info['filename'],
                    'original_path': file_info['path'],
                    'new_path': str(e),
                    'type': file_info['type'],
                    'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'status': 'Failed'
                })
        
        # Save move log
        self.save_move_log(log_entries)
        
        self.progress_label.setText("Done!")
        self.status_label.setText(f"Moved {moved_count} files successfully")
        self.refresh_move_log()
    
    def flag_cleanup(self):
        """Flag files that may need cleanup"""
        if not self.file_list:
            QMessageBox.warning(self, "Warning", "No files found yet. Please scan a folder first.")
            return
        
        self.progress_label.setText("Flagging files...")
        self.status_label.setText("Flagging files that may need cleanup...")
        
        marked_count = 0
        
        for file_info in self.file_list:
            # Add your cleanup logic here
            file_info['marked'] = True  # Mark file for cleanup
            marked_count += 1
        
        self.progress_label.setText("Done!")
        self.status_label.setText(f"Found {marked_count} files that may need cleanup")
        self.refresh_file_list()
    
    def refresh_file_list(self):
        """Refresh the file list display"""
        self.tree.clear()
        
        search_text = self.search_entry.text().lower()
        category_filter = self.category_filter.currentText()
        marked_only = self.marked_only_cb.isChecked()
        
        for file_info in self.file_list:
            # Apply filters
            if search_text and search_text not in file_info['filename'].lower() and search_text not in file_info['path'].lower():
                continue
            if category_filter != "All" and file_info['category'] != category_filter:
                continue
            if marked_only and not file_info.get('marked'):
                continue
            
            # Get program association
            ext = file_info['type']
            programs = self.program_associations.get(ext, 'Unknown')
            if programs == 'Unknown' and ext:
                programs = self.lookup_file_type_online(ext)
            
            item = QTreeWidgetItem(self.tree)
            item.setText(0, file_info['filename'])
            item.setText(1, file_info['path'])
            item.setText(2, f"{file_info['size_kb']:.2f}")
            item.setText(3, file_info['type'])
            item.setText(4, file_info['modified'])
            item.setText(5, file_info['category'])
            item.setText(6, programs)
            item.setText(7, file_info.get('action', ''))
            
            # Color code marked files
            if file_info.get('marked'):
                item.setBackground(0, QColor("#fffacd"))  # Light yellow
    
    def refresh_summary(self):
        """Refresh the category summary"""
        self.summary_tree.clear()
        
        self.category_summary = defaultdict(lambda: {'count': 0, 'size': 0})
        
        for file_info in self.file_list:
            category = file_info['category'] or 'Uncategorized'
            self.category_summary[category]['count'] += 1
            self.category_summary[category]['size'] += file_info['size_kb']
        
        for category, data in self.category_summary.items():
            item = QTreeWidgetItem(self.summary_tree)
            item.setText(0, category)
            item.setText(1, str(data['count']))
            item.setText(2, f"{data['size']:.2f}")
    
    def exit_application(self):
        """Exit the application"""
        if self.file_list:
            reply = QMessageBox.question(self, "Exit", "You have unsaved scan results. Exit anyway?", 
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.No:
                return
        self.close()
    
    def show_help(self):
        """Show help dialog"""
        help_dialog = QDialog(self)
        help_dialog.setWindowTitle("How to Use - FileWhip")
        help_dialog.resize(750, 650)
        
        layout = QVBoxLayout()
        help_text = QTextEdit()
        help_text.setReadOnly(True)
        help_text.setPlainText("""FileWhip v2.0 - Professional File Organizer

HOW TO USE:

1. SELECT SOURCE FOLDER
   - Click "Browse" next to Source Folder
   - Choose the folder you want to organize

2. SCAN & CATEGORIZE
   - Click "Scan & Categorize" button
   - Files will be scanned and automatically categorized by type
   - View results in the File List tab

3. FLAG FOR CLEANUP
   - Click "Flag for Cleanup" to mark files that may need cleanup
   - Marked files will be highlighted in yellow

4. MOVE FILES
   - Select a destination folder
   - Click "Move Marked Files" to move flagged files
   - All moves are logged for undo

5. UNDO MOVES
   - Click "Undo Moves" to reverse recent file moves
   - View move history in the Move Log tab

FEATURES:
- Auto-categorizes 700+ file types
- Scan and save results for later
- Export to CSV/Excel
- Undo file moves
- Track move history

TIPS:
- Use the search bar to find specific files
- Filter by category to see file types
- Save your scan before moving files
- Check the Move Log if something goes wrong
""")
        layout.addWidget(help_text)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(help_dialog.accept)
        layout.addWidget(close_btn)
        
        help_dialog.setLayout(layout)
        help_dialog.exec()
    
    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder to Scan")
        if folder:
            self.folder_entry.setText(folder)
            self.status_label.setText(f"Target: {folder}")
            self.clear_tool_results()
    
    def browse_destination(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Destination Folder")
        if folder:
            self.dest_entry.setText(folder)
            self.status_label.setText(f"Destination: {folder}")
            self.clear_tool_results()
    
    def apply_filter(self):
        """Apply search and filter to file list"""
        self.refresh_file_list()
    
    def save_move_log(self, log_entries):
        """Save move log to file"""
        log_file = os.path.join(os.path.dirname(__file__), "move_log.json")
        try:
            if os.path.exists(log_file):
                with open(log_file, 'r', encoding='utf-8') as f:
                    existing_log = json.load(f)
                log_entries = existing_log + log_entries
            with open(log_file, 'w', encoding='utf-8') as f:
                json.dump(log_entries, f, indent=2)
        except:
            pass
    
    def load_move_log(self):
        """Load move log from file"""
        log_file = os.path.join(os.path.dirname(__file__), "move_log.json")
        if os.path.exists(log_file):
            try:
                with open(log_file, 'r', encoding='utf-8') as f:
                    self.move_log = json.load(f)
                self.refresh_move_log()
            except:
                self.move_log = []
    
    def refresh_move_log(self):
        """Refresh move log display"""
        self.log_tree.clear()
        for entry in self.move_log:
            item = QTreeWidgetItem(self.log_tree)
            item.setText(0, entry.get('filename', ''))
            item.setText(1, entry.get('original_path', ''))
            item.setText(2, entry.get('new_path', ''))
            item.setText(3, entry.get('type', ''))
            item.setText(4, entry.get('timestamp', ''))
            item.setText(5, entry.get('status', ''))
    
    def undo_moves(self):
        """Undo file moves from log"""
        if not self.move_log:
            QMessageBox.information(self, "Info", "No moves to undo.")
            return
        
        # Undo in reverse order
        for entry in reversed(self.move_log):
            if entry['status'] == 'Moved':
                try:
                    if os.path.exists(entry['new_path']):
                        shutil.move(entry['new_path'], entry['original_path'])
                        entry['status'] = 'Undone'
                except Exception as e:
                    print(f"Error undoing move: {e}")
        
        self.save_move_log(self.move_log)
        self.refresh_move_log()
        QMessageBox.information(self, "Success", "Moves have been undone where possible.")
    
    def open_settings(self):
        """Open settings dialog"""
        settings_dialog = QDialog(self)
        settings_dialog.setWindowTitle("Settings")
        settings_dialog.resize(400, 300)
        
        layout = QVBoxLayout()
        
        # Cutoff Date
        layout.addWidget(QLabel("Cutoff Date (YYYY-MM-DD):"))
        cutoff_entry = QLineEdit()
        cutoff_entry.setText(self.config["cutoff_date"])
        layout.addWidget(cutoff_entry)
        
        # Clutter Types
        layout.addWidget(QLabel("Clutter Types (comma-separated):"))
        clutter_entry = QLineEdit()
        clutter_entry.setText(", ".join(self.config["clutter_types"]))
        layout.addWidget(clutter_entry)
        
        # Hash Duplication
        use_hash_var = Qt.CheckState.Unchecked if not self.config["use_hash_duplication"] else Qt.CheckState.Checked
        hash_cb = QCheckBox("Use hash-based duplicate detection (slower but accurate)")
        hash_cb.setCheckState(use_hash_var)
        layout.addWidget(hash_cb)
        
        def save_settings():
            self.config["cutoff_date"] = cutoff_entry.text()
            self.config["clutter_types"] = [x.strip() for x in clutter_entry.text().split(",")]
            self.config["use_hash_duplication"] = hash_cb.isChecked()
            self.save_config()
            settings_dialog.accept()
            QMessageBox.information(self, "Settings", "Settings saved successfully")
        
        save_btn = QPushButton("Save Settings")
        save_btn.clicked.connect(save_settings)
        layout.addWidget(save_btn)
        
        settings_dialog.setLayout(layout)
        settings_dialog.exec()
    
    def show_shortcuts(self):
        """Show keyboard shortcuts dialog"""
        shortcuts_dialog = QDialog(self)
        shortcuts_dialog.setWindowTitle("Keyboard Shortcuts")
        shortcuts_dialog.resize(400, 200)
        
        layout = QVBoxLayout()
        
        shortcuts_text = QTextEdit()
        shortcuts_text.setReadOnly(True)
        shortcuts_text.setPlainText("""Keyboard Shortcuts:

Ctrl+S - Save current scan
Ctrl+O - Load a saved scan  
Ctrl+E - Export to CSV
""")
        layout.addWidget(shortcuts_text)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(shortcuts_dialog.accept)
        layout.addWidget(close_btn)
        
        shortcuts_dialog.setLayout(layout)
        shortcuts_dialog.exec()
    
    def save_config(self):
        """Save configuration to file"""
        config_file = os.path.join(os.path.dirname(__file__), "config.json")
        with open(config_file, 'w') as f:
            json.dump(self.config, f, indent=2)
    
    def get_program_associations(self):
        """Dictionary mapping file extensions to programs that can open them"""
        return {
            # Music
            'mp3': 'Windows Media Player, VLC, iTunes, Winamp',
            'flac': 'VLC, Foobar2000, Winamp',
            'wav': 'Windows Media Player, VLC, Audacity',
            'm4a': 'iTunes, VLC, Windows Media Player',
            'ogg': 'VLC, Winamp, Foobar2000',
            'wma': 'Windows Media Player, VLC',
            # Video
            'mp4': 'VLC, Windows Media Player, QuickTime',
            'avi': 'VLC, Windows Media Player, DivX Player',
            'mkv': 'VLC, MPC-HC, PotPlayer',
            'mov': 'QuickTime, VLC',
            'webm': 'VLC, Chrome, Edge',
            'wmv': 'Windows Media Player, VLC',
            # Documents
            'pdf': 'Adobe Acrobat, Edge, Chrome, Foxit Reader',
            'doc': 'Microsoft Word, LibreOffice Writer',
            'docx': 'Microsoft Word, LibreOffice Writer',
            'odt': 'LibreOffice Writer, Apache OpenOffice',
            'txt': 'Notepad, Notepad++, WordPad',
            'rtf': 'Microsoft Word, WordPad',
            # Spreadsheets
            'xls': 'Microsoft Excel, LibreOffice Calc',
            'xlsx': 'Microsoft Excel, LibreOffice Calc',
            'ods': 'LibreOffice Calc, Apache OpenOffice',
            'csv': 'Excel, Notepad, LibreOffice Calc',
            # Images
            'jpg': 'Windows Photo Viewer, IrfanView, Paint',
            'jpeg': 'Windows Photo Viewer, IrfanView, Paint',
            'png': 'Windows Photo Viewer, IrfanView, Paint',
            'gif': 'Windows Photo Viewer, IrfanView, Paint',
            'bmp': 'Windows Photo Viewer, Paint',
            'svg': 'Chrome, Edge, Inkscape',
            'psd': 'Adobe Photoshop, GIMP',
            # Archives
            'zip': 'Windows Explorer, 7-Zip, WinRAR',
            'rar': 'WinRAR, 7-Zip',
            '7z': '7-Zip, WinRAR',
            'tar': '7-Zip, WinRAR',
            'gz': '7-Zip, WinRAR',
            # Code
            'py': 'Python IDLE, VS Code, PyCharm',
            'js': 'VS Code, Notepad++, Sublime Text',
            'html': 'Chrome, Edge, VS Code',
            'css': 'VS Code, Notepad++, Sublime Text',
            'json': 'VS Code, Notepad++',
            'xml': 'VS Code, Notepad++',
            # Other common
            'exe': 'Windows (executable)',
            'msi': 'Windows Installer',
            'dll': 'Windows (system file)',
        }
    
    def lookup_file_type_online(self, extension):
        """Look up unknown file type online"""
        if extension in self.unknown_type_cache:
            return self.unknown_type_cache[extension]
        
        try:
            # Use filext.com API or similar
            url = f"https://filext.com/file-extension/{extension}"
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                # Simple parsing - in production, use proper HTML parsing
                content = response.text.lower()
                if "description" in content:
                    # Extract description (simplified)
                    result = f"Unknown (found online)"
                    self.unknown_type_cache[extension] = result
                    return result
        except:
            pass
        
        result = "Unknown"
        self.unknown_type_cache[extension] = result
        return result
    
    def extract_audio_metadata(self, file_path):
        """Extract metadata from audio files"""
        if not MUTAGEN_AVAILABLE:
            return None
        
        try:
            audio = MutagenFile(file_path)
            if audio is None:
                return None
            
            metadata = {}
            
            # Try to get common metadata fields
            if hasattr(audio, 'get'):
                # MP3, FLAC, etc.
                metadata['artist'] = audio.get('artist', ['Unknown'])[0] if audio.get('artist') else 'Unknown'
                metadata['album'] = audio.get('album', ['Unknown'])[0] if audio.get('album') else 'Unknown'
                metadata['title'] = audio.get('title', ['Unknown'])[0] if audio.get('title') else 'Unknown'
                metadata['year'] = audio.get('date', ['Unknown'])[0] if audio.get('date') else 'Unknown'
                metadata['genre'] = audio.get('genre', ['Unknown'])[0] if audio.get('genre') else 'Unknown'
                metadata['track'] = audio.get('tracknumber', ['Unknown'])[0] if audio.get('tracknumber') else 'Unknown'
            else:
                # Other formats
                metadata = {'artist': 'Unknown', 'album': 'Unknown', 'title': 'Unknown', 'year': 'Unknown', 'genre': 'Unknown', 'track': 'Unknown'}
            
            return metadata
        except (ID3NoHeaderError, Exception):
            return None
    
    def find_duplicate_files(self):
        """Find duplicate files using multi-stage metadata comparison"""
        if not self.file_list:
            QMessageBox.warning(self, "Warning", "No files to analyze. Scan a folder first.")
            return
        
        # Create dialog to select duplicate detection criteria
        dialog = QDialog(self)
        dialog.setWindowTitle("Duplicate Detection Options")
        dialog.resize(400, 300)
        layout = QVBoxLayout()
        
        layout.addWidget(QLabel("Select duplicate detection criteria:"))
        
        exact_hash_cb = QCheckBox("Exact Hash (100% match)")
        exact_hash_cb.setChecked(True)
        layout.addWidget(exact_hash_cb)
        
        same_name_cb = QCheckBox("Same Filename (potential duplicates)")
        same_name_cb.setChecked(True)
        layout.addWidget(same_name_cb)
        
        same_size_cb = QCheckBox("Same Size (potential duplicates)")
        same_size_cb.setChecked(True)
        layout.addWidget(same_size_cb)
        
        same_date_cb = QCheckBox("Same Modification Date (recent copies)")
        same_date_cb.setChecked(False)
        layout.addWidget(same_date_cb)
        
        similar_name_cb = QCheckBox("Similar Filename (fuzzy match)")
        similar_name_cb.setChecked(False)
        layout.addWidget(similar_name_cb)
        
        def start_detection():
            criteria = {
                'exact_hash': exact_hash_cb.isChecked(),
                'same_name': same_name_cb.isChecked(),
                'same_size': same_size_cb.isChecked(),
                'same_date': same_date_cb.isChecked(),
                'similar_name': similar_name_cb.isChecked()
            }
            dialog.accept()
            self.run_duplicate_detection(criteria)
        
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(start_detection)
        btn_box.rejected.connect(dialog.reject)
        layout.addWidget(btn_box)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def run_duplicate_detection(self, criteria):
        """Run duplicate detection with selected criteria (AND logic)"""
        self.status_label.setText("Analyzing files...")
        self.progress_bar.setValue(0)
        
        # Build a dictionary to group files by their matching criteria
        # For hash-based duplicates, group by hash
        duplicate_groups = []
        duplicate_count = 0
        
        if criteria['exact_hash']:
            # Pre-calculate all hashes
            hash_dict = defaultdict(list)
            for file_info in self.file_list:
                file_hash = self.calculate_file_hash(file_info['path'])
                if file_hash:
                    hash_dict[file_hash].append(file_info)
            
            # Find groups with multiple files
            for file_hash, files in hash_dict.items():
                if len(files) > 1:
                    duplicate_groups.append({
                        'files': files,
                        'reason': 'Exact Hash'
                    })
                    for i, file_info in enumerate(files[1:], 1):
                        file_info['marked'] = True
                        file_info['action'] = f"Duplicate #{i+1} (Exact Hash)"
                        duplicate_count += 1
        
        # For other criteria, group by combined key
        if criteria['same_name'] or criteria['same_size'] or criteria['same_date'] or criteria['similar_name']:
            key_dict = defaultdict(list)
            for file_info in self.file_list:
                # Create a key based on selected criteria
                key_parts = []
                if criteria['same_name']:
                    key_parts.append(file_info['filename'])
                if criteria['same_size']:
                    key_parts.append(str(int(file_info['size_kb'])))
                if criteria['same_date']:
                    key_parts.append(file_info['modified'])
                if criteria['similar_name']:
                    # For similar names, we need to compare pairs, not group by key
                    pass
                
                if key_parts:
                    key = "|".join(key_parts)
                    key_dict[key].append(file_info)
            
            # Find groups with multiple files
            for key, files in key_dict.items():
                if len(files) > 1:
                    reasons = []
                    if criteria['same_name']:
                        reasons.append("Same Name")
                    if criteria['same_size']:
                        reasons.append("Same Size")
                    if criteria['same_date']:
                        reasons.append("Same Date")
                    
                    duplicate_groups.append({
                        'files': files,
                        'reason': ', '.join(reasons)
                    })
                    for i, file_info in enumerate(files[1:], 1):
                        file_info['marked'] = True
                        file_info['action'] = f"Duplicate #{i+1} ({', '.join(reasons)})"
                        duplicate_count += 1
        
        # Handle similar name matching separately (pairwise comparison)
        if criteria['similar_name'] and not (criteria['same_name'] or criteria['same_size'] or criteria['same_date']):
            processed_indices = set()
            for i, file_info in enumerate(self.file_list):
                if i in processed_indices:
                    continue
                similar_files = [file_info]
                for j in range(i + 1, len(self.file_list)):
                    if j in processed_indices:
                        continue
                    if self.similar_strings(file_info['filename'], self.file_list[j]['filename'], threshold=0.8):
                        similar_files.append(self.file_list[j])
                        processed_indices.add(j)
                
                if len(similar_files) > 1:
                    duplicate_groups.append({
                        'files': similar_files,
                        'reason': 'Similar Name'
                    })
                    for i, file_info in enumerate(similar_files[1:], 1):
                        file_info['marked'] = True
                        file_info['action'] = f"Duplicate #{i+1} (Similar Name)"
                        duplicate_count += 1
                    processed_indices.add(i)
        
        self.progress_bar.setValue(100)
        self.status_label.setText("Duplicate check complete")
        self.refresh_file_list()
        
        # Display results in Tool Results tab
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        criteria_str = ', '.join([k for k, v in criteria.items() if v])
        
        if duplicate_groups:
            for group in duplicate_groups:
                files = group['files']
                reason = group['reason']
                # Show all locations for this duplicate group
                locations = " | ".join([os.path.dirname(f['path']) for f in files])
                item = QTreeWidgetItem(self.results_tree)
                item.setText(0, "Duplicate Detection")
                item.setText(1, files[0]['filename'])
                item.setText(2, f"Duplicate group ({len(files)} files) | Reason: {reason} | Size: {files[0]['size_kb']:.2f} KB | All locations: {locations}")
                item.setText(3, timestamp)
        else:
            item = QTreeWidgetItem(self.results_tree)
            item.setText(0, "Duplicate Detection")
            item.setText(1, "No duplicates found")
            item.setText(2, f"Criteria: {criteria_str}")
            item.setText(3, timestamp)
        
        self.tab_widget.setCurrentWidget(self.results_tab)
    
    def calculate_file_hash(self, file_path):
        """Calculate MD5 hash of a file"""
        try:
            if os.path.exists(file_path):
                hash_md5 = hashlib.md5()
                with open(file_path, 'rb') as f:
                    while chunk := f.read(8192):
                        hash_md5.update(chunk)
                return hash_md5.hexdigest()
        except:
            pass
        return None
    
    def similar_strings(self, s1, s2, threshold=0.8):
        """Calculate similarity between two strings using Levenshtein distance"""
        if not s1 or not s2:
            return False
        
        # Simple similarity check based on common substrings
        s1_lower = s1.lower()
        s2_lower = s2.lower()
        
        # Check if one is substring of the other
        if s1_lower in s2_lower or s2_lower in s1_lower:
            return True
        
        # Check for common prefix/suffix
        common_prefix = 0
        for a, b in zip(s1_lower, s2_lower):
            if a == b:
                common_prefix += 1
            else:
                break
        
        common_suffix = 0
        for a, b in zip(reversed(s1_lower), reversed(s2_lower)):
            if a == b:
                common_suffix += 1
            else:
                break
        
        # Calculate similarity based on common characters
        max_len = max(len(s1), len(s2))
        if max_len == 0:
            return True
        
        similarity = (common_prefix + common_suffix) / max_len
        return similarity >= threshold
    
    def find_large_files(self, size_mb=100):
        """Find files larger than specified size in MB"""
        if not self.file_list:
            QMessageBox.warning(self, "Warning", "No files to analyze. Scan a folder first.")
            return
        
        large_files = [f for f in self.file_list if f['size_kb'] > size_mb * 1024]
        
        for file_info in large_files:
            file_info['marked'] = True
            file_info['action'] = f"Large file ({file_info['size_kb']/1024:.1f} MB)"
        
        self.refresh_file_list()
        
        # Display results in Tool Results tab
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if large_files:
            for file_info in large_files:
                item = QTreeWidgetItem(self.results_tree)
                item.setText(0, "Large Files")
                item.setText(1, file_info['filename'])
                item.setText(2, f"Size: {file_info['size_kb']/1024:.2f} MB | Path: {file_info['path']}")
                item.setText(3, timestamp)
        else:
            item = QTreeWidgetItem(self.results_tree)
            item.setText(0, "Large Files")
            item.setText(1, "No large files found")
            item.setText(2, f"Threshold: {size_mb} MB")
            item.setText(3, timestamp)
        
        self.tab_widget.setCurrentWidget(self.results_tab)
    
    def find_empty_folders(self):
        """Find empty folders in the scanned directory"""
        target_folder = self.folder_entry.text()
        if not target_folder or not os.path.exists(target_folder):
            QMessageBox.warning(self, "Warning", "Please select a valid folder first.")
            return
        
        empty_folders = []
        for root, dirs, files in os.walk(target_folder):
            for dir_name in dirs:
                dir_path = os.path.join(root, dir_name)
                if not os.listdir(dir_path):  # Empty folder
                    empty_folders.append(dir_path)
        
        # Display results in Tool Results tab
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if empty_folders:
            for folder_path in empty_folders[:100]:  # Show first 100
                item = QTreeWidgetItem(self.results_tree)
                item.setText(0, "Empty Folders")
                item.setText(1, os.path.basename(folder_path))
                item.setText(2, folder_path)
                item.setText(3, timestamp)
            if len(empty_folders) > 100:
                item = QTreeWidgetItem(self.results_tree)
                item.setText(0, "Empty Folders")
                item.setText(1, f"... and {len(empty_folders) - 100} more")
                item.setText(2, "")
                item.setText(3, timestamp)
        else:
            item = QTreeWidgetItem(self.results_tree)
            item.setText(0, "Empty Folders")
            item.setText(1, "No empty folders found")
            item.setText(2, f"Target: {target_folder}")
            item.setText(3, timestamp)
        
        self.tab_widget.setCurrentWidget(self.results_tab)
    
    def batch_rename_files(self):
        """Batch rename files with pattern replacement"""
        if not self.file_list:
            QMessageBox.warning(self, "Warning", "No files to rename. Scan a folder first.")
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Batch Rename")
        dialog.resize(400, 200)
        
        layout = QVBoxLayout()
        
        layout.addWidget(QLabel("Find:"))
        find_entry = QLineEdit()
        layout.addWidget(find_entry)
        
        layout.addWidget(QLabel("Replace with:"))
        replace_entry = QLineEdit()
        layout.addWidget(replace_entry)
        
        def execute_rename():
            find_text = find_entry.text()
            replace_text = replace_entry.text()
            
            if not find_text:
                QMessageBox.warning(dialog, "Warning", "Please enter text to find.")
                return
            
            renamed_count = 0
            for file_info in self.file_list:
                if file_info.get('marked'):
                    old_path = file_info['path']
                    old_name = file_info['filename']
                    new_name = old_name.replace(find_text, replace_text)
                    
                    if new_name != old_name:
                        new_path = os.path.join(os.path.dirname(old_path), new_name)
                        try:
                            os.rename(old_path, new_path)
                            renamed_count += 1
                        except Exception as e:
                            QMessageBox.warning(dialog, "Error", f"Failed to rename {old_name}: {e}")
            
            # Display results in Tool Results tab
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if renamed_count > 0:
                for file_info in self.file_list:
                    if file_info.get('marked'):
                        old_name = file_info['filename']
                        new_name = old_name.replace(find_text, replace_text)
                        if new_name != old_name:
                            item = QTreeWidgetItem(self.results_tree)
                            item.setText(0, "Batch Rename")
                            item.setText(1, old_name)
                            item.setText(2, f"Renamed to: {new_name} | Path: {file_info['path']}")
                            item.setText(3, timestamp)
            else:
                item = QTreeWidgetItem(self.results_tree)
                item.setText(0, "Batch Rename")
                item.setText(1, "No files renamed")
                item.setText(2, f"Pattern: {find_text} -> {replace_text}")
                item.setText(3, timestamp)
            
            self.tab_widget.setCurrentWidget(self.results_tab)
            
            QMessageBox.information(dialog, "Success", f"Renamed {renamed_count} files")
            dialog.accept()
        
        rename_btn = QPushButton("Execute Rename")
        rename_btn.clicked.connect(execute_rename)
        layout.addWidget(rename_btn)
        
        dialog.setLayout(layout)
        dialog.exec()
    
    def organize_music_by_metadata(self):
        """Organize music files by artist/album using metadata"""
        if not MUTAGEN_AVAILABLE:
            QMessageBox.warning(self, "Warning", "Mutagen library not installed. Cannot extract audio metadata.")
            return
        
        if not self.file_list:
            QMessageBox.warning(self, "Warning", "No files to organize. Scan a folder first.")
            return
        
        # Filter only music files
        music_files = [f for f in self.file_list if f['category'] == 'Music']
        
        if not music_files:
            QMessageBox.warning(self, "Warning", "No music files found in the scan.")
            return
        
        organized_count = 0
        for file_info in music_files:
            metadata = self.extract_audio_metadata(file_info['path'])
            if metadata:
                artist = metadata['artist'].replace('/', '_').replace('\\', '_')
                album = metadata['album'].replace('/', '_').replace('\\', '_')
                
                if artist != 'Unknown' and album != 'Unknown':
                    # Create folder structure: Artist/Album/
                    dest_folder = self.dest_entry.text()
                    if not dest_folder:
                        QMessageBox.warning(self, "Warning", "Please select a destination folder first.")
                        return
                    
                    target_dir = os.path.join(dest_folder, artist, album)
                    os.makedirs(target_dir, exist_ok=True)
                    
                    # Move file
                    src = file_info['path']
                    dest = os.path.join(target_dir, file_info['filename'])
                    
                    try:
                        if not os.path.exists(dest):
                            shutil.move(src, dest)
                            organized_count += 1
                            file_info['action'] = f"Organized: {artist}/{album}"
                    except Exception as e:
                        print(f"Error moving {file_info['filename']}: {e}")
        
        self.refresh_file_list()
        
        # Display results in Tool Results tab
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        if organized_count > 0:
            for file_info in music_files:
                metadata = self.extract_audio_metadata(file_info['path'])
                if metadata:
                    artist = metadata['artist'].replace('/', '_').replace('\\', '_')
                    album = metadata['album'].replace('/', '_').replace('\\', '_')
                    if artist != 'Unknown' and album != 'Unknown':
                        item = QTreeWidgetItem(self.results_tree)
                        item.setText(0, "Music Organization")
                        item.setText(1, file_info['filename'])
                        item.setText(2, f"Artist: {artist} | Album: {album} | Original: {file_info['path']}")
                        item.setText(3, timestamp)
        else:
            item = QTreeWidgetItem(self.results_tree)
            item.setText(0, "Music Organization")
            item.setText(1, "No files organized")
            item.setText(2, "Missing metadata or already organized")
            item.setText(3, timestamp)
        
        self.tab_widget.setCurrentWidget(self.results_tab)
        
        QMessageBox.information(self, "Music Organized", f"Organized {organized_count} music files by artist/album")
    
    def get_known_file_types(self):
        """Dictionary mapping file extensions to categories"""
        return {
            # Music
            'aac': 'Music', 'aiff': 'Music', 'alac': 'Music', 'flac': 'Music',
            'mod': 'Music', 'mp3': 'Music', 'm4a': 'Music', 'ogg': 'Music',
            'snd': 'Music', 'wav': 'Music', 'xm': 'Music', 'wma': 'Music',
            'amr': 'Music', 'aup': 'Music Project', 'ra': 'Music',
            'mid': 'Music', 'ptb': 'Music', 'it': 'Music', 's3m': 'Music',
            '669': 'Music', 'med': 'Music', 'mtm': 'Music', 'psm': 'Music',
            'umx': 'Music', 'gym': 'Music', 'dts': 'Music', 'ac3': 'Music',
            'mka': 'Music', 'm4r': 'Music', 'aif': 'Music', 'aifc': 'Music',
            'gsm': 'Music', 'au': 'Music', 'snd': 'Music', 'ulaw': 'Music',
            'vox': 'Music', 'dwd': 'Music', 'smv': 'Music', 'vqf': 'Music',
            'swa': 'Music', 'cda': 'Music', 'act': 'Music', 'au3': 'Music',
            'oma': 'Music', 'atrac': 'Music', 'aa': 'Music', 'm4b': 'Music',
            'm4p': 'Music', 'mpc': 'Music', 'mp+': 'Music', 'mpp': 'Music',
            'rm': 'Music', 'ram': 'Music', 'ra': 'Music', 'wv': 'Music',
            'shn': 'Music', 'ape': 'Music', 'ofr': 'Music', 'ofs': 'Music',
            'tak': 'Music', 'tta': 'Music', 'mlp': 'Music', 'dts': 'Music',
            # Video
            '3gp': 'Video', 'avi': 'Video', 'mkv': 'Video', 'mov': 'Video',
            'mp4': 'Video', 'mpeg': 'Video', 'mpg': 'Video', 'mts': 'Video',
            'vob': 'Video', 'webm': 'Video', 'wmv': 'Video', 'flv': 'Video',
            'rmvb': 'Video', 'asf': 'Video', 'asx': 'Video', 'm4v': 'Video',
            'f4v': 'Video', 'f4p': 'Video', 'f4a': 'Video', '3g2': 'Video',
            '3gp2': 'Video', '3gpp': 'Video', 'divx': 'Video', 'xvid': 'Video',
            'ts': 'Video', 'm2ts': 'Video', 'm2t': 'Video', 'mts': 'Video',
            'mxf': 'Video', 'dv': 'Video', 'dvi': 'Video', 'ogv': 'Video',
            'ogx': 'Video', 'qt': 'Video', 'moov': 'Video', 'rm': 'Video',
            'rv': 'Video', 'yuv': 'Video', 'y4m': 'Video', 'mjpeg': 'Video',
            'mj2': 'Video', 'mjp': 'Video', 'jpm': 'Video', 'jpgv': 'Video',
            'mpg2': 'Video', 'mp2v': 'Video', 'mpe': 'Video', 'mpv': 'Video',
            'm1v': 'Video', 'm2v': 'Video', 'mpv2': 'Video', 'mp2': 'Video',
            'svi': 'Video', '3iv': 'Video', 'amv': 'Video', 'nsv': 'Video',
            'flc': 'Video', 'fli': 'Video', 'afl': 'Video', 'rpl': 'Video',
            'wmx': 'Video', 'wvx': 'Video', 'wax': 'Video', 'wmp': 'Video',
            'scm': 'Video', 'viv': 'Video', 'vivo': 'Video', 'ivf': 'Video',
            'cpi': 'Video', 'dvx': 'Video', 'gvi': 'Video', 'gvp': 'Video',
            'hdmov': 'Video', 'k3g': 'Video', 'm15': 'Video', 'm4e': 'Video',
            'm75': 'Video', 'mp21': 'Video', 'mp4v': 'Video', 'mpg4': 'Video',
            'mqv': 'Video', 'sdv': 'Video', 'sdp': 'Video', 'smi': 'Video',
            'ssm': 'Video', 'trp': 'Video', 'ty': 'Video', 'uvf': 'Video',
            'uvu': 'Video', 'uvvp': 'Video', 'uvvs': 'Video', 'vob': 'Video',
            'vp6': 'Video', 'vp7': 'Video', 'webm': 'Video', 'wm': 'Video',
            'wmx': 'Video', 'wvx': 'Video', 'xmv': 'Video', 'xsv': 'Video',
            # Spreadsheets
            'csv': 'Spreadsheet', 'ods': 'Spreadsheet', 'xls': 'Spreadsheet',
            'xlsb': 'Spreadsheet', 'xlsx': 'Spreadsheet', 'xlsm': 'Spreadsheet',
            'numbers': 'Spreadsheet', 'sxc': 'Spreadsheet', 'stc': 'Spreadsheet',
            'dif': 'Spreadsheet', 'dbf': 'Spreadsheet', 'wks': 'Spreadsheet',
            'wk1': 'Spreadsheet', 'wk3': 'Spreadsheet', 'wk4': 'Spreadsheet',
            'wq1': 'Spreadsheet', 'wq2': 'Spreadsheet', '123': 'Spreadsheet',
            'slk': 'Spreadsheet', 'pxl': 'Spreadsheet', 'gnumeric': 'Spreadsheet',
            'ots': 'Spreadsheet', 'fods': 'Spreadsheet', 'xlr': 'Spreadsheet',
            # Documents
            'doc': 'Document', 'docx': 'Document', 'djvu': 'Document',
            'epub': 'Document', 'md': 'Document', 'mobi': 'Document',
            'odt': 'Document', 'pdf': 'Document', 'rst': 'Document',
            'rtf': 'Document', 'tex': 'Document', 'txt': 'Document',
            'docm': 'Document', 'dot': 'Document', 'dotm': 'Document',
            'dotx': 'Document', 'odm': 'Document', 'ott': 'Document',
            'sxw': 'Document', 'stw': 'Document', 'wpd': 'Document',
            'wps': 'Document', 'wpt': 'Document', 'abw': 'Document',
            'zabw': 'Document', 'lwp': 'Document', 'pages': 'Document',
            'cwk': 'Document', 'mcw': 'Document', 'pdb': 'Document',
            'pdx': 'Document', 'xps': 'Document', 'oxps': 'Document',
            'ps': 'Document', 'eps': 'Document', 'prn': 'Document',
            'wp': 'Document', 'wri': 'Document', 'hlp': 'Document',
            'man': 'Document', 'info': 'Document', '1st': 'Document',
            # Images
            'ai': 'Image', 'apng': 'Image', 'bmp': 'Image', 'exr': 'Image',
            'gif': 'Image', 'heic': 'Image', 'ico': 'Image', 'indd': 'Image',
            'jpg': 'Image', 'jpeg': 'Image', 'png': 'Image', 'psd': 'Image',
            'raw': 'Image', 'svg': 'Image', 'tiff': 'Image', 'webp': 'Image',
            'xcf': 'Image', 'cr2': 'Image', 'crw': 'Image', 'dcr': 'Image',
            'dng': 'Image', 'erf': 'Image', 'k25': 'Image', 'kdc': 'Image',
            'mrw': 'Image', 'nef': 'Image', 'orf': 'Image', 'pef': 'Image',
            'raf': 'Image', 'raw': 'Image', 'rw2': 'Image', 'srf': 'Image',
            'sr2': 'Image', 'arw': 'Image', '3fr': 'Image', 'fff': 'Image',
            'iiq': 'Image', 'mos': 'Image', 'nrw': 'Image', 'ptx': 'Image',
            'pxn': 'Image', 'r3d': 'Image', 'srw': 'Image', 'x3f': 'Image',
            'tif': 'Image', 'tga': 'Image', 'pcx': 'Image', 'pic': 'Image',
            'pct': 'Image', 'pnm': 'Image', 'pbm': 'Image', 'pgm': 'Image',
            'ppm': 'Image', 'pam': 'Image', 'pfm': 'Image', 'sgi': 'Image',
            'img': 'Image', 'jpe': 'Image', 'jfif': 'Image', 'jif': 'Image',
            'jp2': 'Image', 'j2k': 'Image', 'jpf': 'Image', 'jpx': 'Image',
            'jpm': 'Image', 'mj2': 'Image', 'bpg': 'Image', 'heif': 'Image',
            'hdp': 'Image', 'avif': 'Image', 'jxl': 'Image', 'qoi': 'Image',
            'pvr': 'Image', 'pvrtc': 'Image', 'pvr.gz': 'Image', 'ktx': 'Image',
            'dds': 'Image', 'astc': 'Image', 'pkm': 'Image', 'pvr': 'Image',
            'vicar': 'Image', 'fits': 'Image', 'ipl': 'Image', 'ndpi': 'Image',
            'scn': 'Image', 'scan': 'Image', 'dcx': 'Image', 'epsf': 'Image',
            'epsi': 'Image', 'eps2': 'Image', 'eps3': 'Image', 'epsi': 'Image',
            'psd1': 'Image', 'psb': 'Image', 'pdd': 'Image', 'psdt': 'Image',
            'cpt': 'Image', 'csh': 'Image', 'icl': 'Image', 'cur': 'Image',
            'ani': 'Image', 'icns': 'Image', 'icon': 'Image', 'icr': 'Image',
            'rsb': 'Image', 'rsrc': 'Image', 'rlc': 'Image', 'dib': 'Image',
            'rle': 'Image', 'wdp': 'Image', 'hdp': 'Image', 'jxr': 'Image',
            'wmp': 'Image', 'blp': 'Image', 'vtf': 'Image', 'tex': 'Image',
            'pvr': 'Image', 'dds': 'Image', 'ftc': 'Image', 'ftu': 'Image',
            # Presentations
            'key': 'Presentation', 'odp': 'Presentation', 'ppt': 'Presentation',
            'pptx': 'Presentation', 'pptm': 'Presentation', 'sxi': 'Presentation',
            'pps': 'Presentation', 'ppsx': 'Presentation', 'ppsm': 'Presentation',
            'pot': 'Presentation', 'potx': 'Presentation', 'potm': 'Presentation',
            'ppa': 'Presentation', 'ppam': 'Presentation', 'dps': 'Presentation',
            'dpt': 'Presentation', 'shw': 'Presentation', 'show': 'Presentation',
            'sldx': 'Presentation', 'sldm': 'Presentation', 'thmx': 'Presentation',
            'prz': 'Presentation', 'uop': 'Presentation', 'uopx': 'Presentation',
            'uotp': 'Presentation', 'uot': 'Presentation', 'uotp': 'Presentation',
            # Web
            'css': 'Web', 'html': 'Web', 'htm': 'Web', 'xhtml': 'Web',
            'xht': 'Web', 'mht': 'Web', 'mhtml': 'Web', 'shtml': 'Web',
            'phtml': 'Web', 'jsp': 'Web', 'jspx': 'Web', 'php': 'Web',
            'php3': 'Web', 'php4': 'Web', 'php5': 'Web', 'phtml': 'Web',
            'asp': 'Web', 'aspx': 'Web', 'asax': 'Web', 'ashx': 'Web',
            'asmx': 'Web', 'axd': 'Web', 'svc': 'Web', 'js': 'Web',
            'json': 'Web', 'xml': 'Web', 'xaml': 'Web', 'xsl': 'Web',
            'xslt': 'Web', 'yaml': 'Web', 'yml': 'Web', 'ini': 'Web',
            'conf': 'Web', 'config': 'Web', 'htaccess': 'Web', 'htpasswd': 'Web',
            'ts': 'Web', 'tsx': 'Web', 'jsx': 'Web', 'vue': 'Web',
            'svelte': 'Web', 'graphql': 'Web', 'gql': 'Web', 'mustache': 'Web',
            'hbs': 'Web', 'handlebars': 'Web', 'erb': 'Web', 'ejs': 'Web',
            'pug': 'Web', 'jade': 'Web', 'haml': 'Web', 'slim': 'Web',
            'less': 'Web', 'scss': 'Web', 'sass': 'Web', 'styl': 'Web',
            'stylus': 'Web', 'coffee': 'Web', 'litcoffee': 'Web',
            'tsv': 'Web', 'rss': 'Web', 'atom': 'Web', 'opml': 'Web',
            'sgml': 'Web', 'wml': 'Web', 'wmls': 'Web', 'svg': 'Web',
            'svgz': 'Web', 'swf': 'Web', 'swfl': 'Web', 'fla': 'Web',
            'as': 'Web', 'mxp': 'Web', 'swc': 'Web', 'air': 'Web',
            'crx': 'Web', 'xpi': 'Web', 'jar': 'Web', 'war': 'Web',
            'ear': 'Web', 'nar': 'Web', 'apk': 'Web', 'ipa': 'Web',
            'deb': 'Web', 'rpm': 'Web', 'dmg': 'Web', 'pkg': 'Web',
            'msi': 'Web', 'exe': 'Web', 'app': 'Web', 'appimage': 'Web',
            'snap': 'Web', 'flatpak': 'Web',
            # Config/System
            'cfg': 'Config', 'dll': 'System', 'env': 'Config',
            'log': 'Log', 'ocx': 'System', 'plist': 'Config',
            'sys': 'System', 'tmp': 'Temporary', 'bak': 'Backup',
            'old': 'Backup', 'dmp': 'System', 'lock': 'System',
            'pid': 'System', 'dat': 'Data', 'db': 'Database',
            'sqlite': 'Database', 'sqlite3': 'Database', 'mdb': 'Database',
            'accdb': 'Database', 'dbf': 'Database', 'db3': 'Database',
            'db-shm': 'Database', 'db-wal': 'Database', 'sdb': 'Database',
            'sdf': 'Database', 'mdf': 'Database', 'ndf': 'Database',
            'ldf': 'Database', 'ibd': 'Database', 'myd': 'Database',
            'myi': 'Database', 'frm': 'Database', 'sql': 'Database',
            'sql.gz': 'Database', 'sql.bz2': 'Database', 'sql.tar.gz': 'Database',
            'dump': 'Database', 'backup': 'Backup', 'sqlc': 'Database',
            'sqlite-shm': 'Database', 'sqlite-wal': 'Database',
            'properties': 'Config', 'prefs': 'Config', 'setting': 'Config',
            'settings': 'Config', 'conf': 'Config', 'config': 'Config',
            'xml': 'Config', 'json': 'Config', 'yaml': 'Config',
            'yml': 'Config', 'toml': 'Config', 'ini': 'Config',
            'reg': 'Config', 'inf': 'Config', 'manifest': 'Config',
            'desktop': 'Config', 'service': 'Config', 'mount': 'Config',
            'fstab': 'Config', 'hosts': 'Config', 'profile': 'Config',
            'bashrc': 'Config', 'bash_profile': 'Config', 'zshrc': 'Config',
            'vimrc': 'Config', 'gvimrc': 'Config', 'editorconfig': 'Config',
            'gitconfig': 'Config', 'gitignore': 'Config', 'gitattributes': 'Config',
            'gitmodules': 'Config', 'gitkeep': 'Config', 'makefile': 'Code',
            'make': 'Code', 'cmake': 'Code', 'dockerfile': 'Code',
            'dockerignore': 'Config', 'containerfile': 'Code',
            'procfile': 'Config', 'rakefile': 'Code', 'gemfile': 'Code',
            'gemspec': 'Code', 'podspec': 'Code', 'cartfile': 'Code',
            'podfile': 'Code', 'composer.json': 'Config', 'composer.lock': 'Config',
            'package.json': 'Config', 'package-lock.json': 'Config',
            'yarn.lock': 'Config', 'pnpm-lock.yaml': 'Config',
            'tsconfig.json': 'Config', 'jsconfig.json': 'Config',
            'babelrc': 'Config', 'eslintrc': 'Config', 'eslintrc.json': 'Config',
            'eslintrc.js': 'Config', 'eslintrc.yaml': 'Config',
            'eslintrc.yml': 'Config', 'prettierrc': 'Config',
            'prettierrc.json': 'Config', 'prettierrc.js': 'Config',
            'prettierrc.yaml': 'Config', 'prettierrc.yml': 'Config',
            'stylelintrc': 'Config', 'stylelintrc.json': 'Config',
            'stylelintrc.js': 'Config', 'stylelintrc.yaml': 'Config',
            'stylelintrc.yml': 'Config', 'editorconfig': 'Config',
            'tslint.json': 'Config', 'pylintrc': 'Config',
            'flake8': 'Config', 'pycodestyle': 'Config', 'pydocstyle': 'Config',
            'pyflakes': 'Config', 'mypy.ini': 'Config', 'setup.cfg': 'Config',
            'setup.py': 'Code', 'requirements.txt': 'Config',
            'requirements-dev.txt': 'Config', 'pipfile': 'Config',
            'pipfile.lock': 'Config', 'poetry.lock': 'Config',
            'pyproject.toml': 'Config', 'tox.ini': 'Config',
            'pytest.ini': 'Config', 'conftest.py': 'Code',
            'coveragerc': 'Config', '.coveragerc': 'Config',
            'py.typed': 'Config', 'MANIFEST.in': 'Config',
            'CHANGELOG': 'Document', 'CHANGELOG.md': 'Document',
            'CHANGES': 'Document', 'CHANGES.md': 'Document',
            'LICENSE': 'Document', 'LICENSE.txt': 'Document',
            'LICENSE.md': 'Document', 'COPYING': 'Document',
            'README': 'Document', 'README.md': 'Document',
            'README.txt': 'Document', 'AUTHORS': 'Document',
            'CONTRIBUTORS': 'Document', 'INSTALL': 'Document',
            'NEWS': 'Document', 'HISTORY': 'Document',
            'VERSION': 'Document', 'CREDITS': 'Document',
            'FAQ': 'Document', 'TODO': 'Document', 'THANKS': 'Document',
            # Archives
            'zip': 'Archive', 'rar': 'Archive', '7z': 'Archive',
            'tar': 'Archive', 'gz': 'Archive', 'bz2': 'Archive',
            'xz': 'Archive', 'lzma': 'Archive', 'cab': 'Archive',
            'iso': 'Archive', 'img': 'Archive', 'dmg': 'Archive',
            'udf': 'Archive', 'toast': 'Archive', 'pkg': 'Archive',
            'msi': 'Archive', 'deb': 'Archive', 'rpm': 'Archive',
            'apk': 'Archive', 'ipa': 'Archive', 'app': 'Archive',
            'appimage': 'Archive', 'snap': 'Archive', 'flatpak': 'Archive',
            'tgz': 'Archive', 'tbz2': 'Archive', 'txz': 'Archive',
            'tlzma': 'Archive', 'tar.gz': 'Archive', 'tar.bz2': 'Archive',
            'tar.xz': 'Archive', 'tar.lzma': 'Archive', 'tar.Z': 'Archive',
            'taz': 'Archive', 'arj': 'Archive', 'ace': 'Archive',
            'bzip2': 'Archive', 'gz': 'Archive', 'lha': 'Archive',
            'lzh': 'Archive', 'lzx': 'Archive', 'pak': 'Archive',
            'sea': 'Archive', 'sit': 'Archive', 'sitx': 'Archive',
            'sqx': 'Archive', 'uc': 'Archive', 'uha': 'Archive',
            'war': 'Archive', 'ear': 'Archive', 'nar': 'Archive',
            'jar': 'Archive', 'sar': 'Archive', 'par': 'Archive',
            'cbr': 'Archive', 'cbz': 'Archive', 'cb7': 'Archive',
            'cba': 'Archive', 'epub': 'Archive', 'mobi': 'Archive',
            'azw': 'Archive', 'azw3': 'Archive', 'kfx': 'Archive',
            'kf8': 'Archive', 'lit': 'Archive', 'pdb': 'Archive',
            'prc': 'Archive', 'tcr': 'Archive', 'xps': 'Archive',
            'oxps': 'Archive', 'chm': 'Archive', 'chw': 'Archive',
            'hlp': 'Archive', 'ehelp': 'Archive',
            # Code
            'py': 'Code', 'pyw': 'Code', 'pyc': 'Code',
            'pyo': 'Code', 'pyd': 'Code', 'pyx': 'Code',
            'pyi': 'Code', 'rpy': 'Code', 'py3': 'Code',
            'c': 'Code', 'cpp': 'Code', 'cc': 'Code',
            'cxx': 'Code', 'h': 'Code', 'hpp': 'Code',
            'hh': 'Code', 'hxx': 'Code', 'h++': 'Code',
            'm': 'Code', 'mm': 'Code', 'swift': 'Code',
            'java': 'Code', 'jar': 'Code', 'war': 'Code',
            'class': 'Code', 'js': 'Code', 'jsx': 'Code',
            'ts': 'Code', 'tsx': 'Code', 'vue': 'Code',
            'rb': 'Code', 'go': 'Code', 'rs': 'Code',
            'php': 'Code', 'php3': 'Code', 'php4': 'Code',
            'php5': 'Code', 'phtml': 'Code', 'asp': 'Code',
            'aspx': 'Code', 'cs': 'Code', 'vb': 'Code',
            'vbproj': 'Code', 'csproj': 'Code', 'sln': 'Code',
            'pl': 'Code', 'pm': 'Code', 't': 'Code',
            'sh': 'Code', 'bash': 'Code', 'zsh': 'Code',
            'fish': 'Code', 'csh': 'Code', 'tcsh': 'Code',
            'ksh': 'Code', 'awk': 'Code', 'sed': 'Code',
            'lua': 'Code', 'tcl': 'Code', 'sql': 'Code',
            'r': 'Code', 'rmd': 'Code', 'jl': 'Code',
            'scala': 'Code', 'sc': 'Code', 'groovy': 'Code',
            'gvy': 'Code', 'gy': 'Code', 'gsh': 'Code',
            'kt': 'Code', 'kts': 'Code', 'kotlin': 'Code',
            'dart': 'Code', 'nim': 'Code', 'nims': 'Code',
            'cr': 'Code', 'd': 'Code', 'di': 'Code',
            'f90': 'Code', 'f95': 'Code', 'f03': 'Code',
            'f08': 'Code', 'f': 'Code', 'for': 'Code',
            'cob': 'Code', 'cbl': 'Code', 'cpy': 'Code',
            'lisp': 'Code', 'lsp': 'Code', 'cl': 'Code',
            'el': 'Code', 'scm': 'Code', 'ss': 'Code',
            'rkt': 'Code', 'ml': 'Code', 'mli': 'Code',
            'hs': 'Code', 'lhs': 'Code', 'erl': 'Code',
            'hrl': 'Code', 'ex': 'Code', 'exs': 'Code',
            'fs': 'Code', 'fsi': 'Code', 'fsx': 'Code',
            'fsscript': 'Code', 'v': 'Code', 'sv': 'Code',
            'verilog': 'Code', 'vhdl': 'Code', 'vhd': 'Code',
            'asm': 'Code', 's': 'Code', 'a51': 'Code',
            'inc': 'Code', 'nasm': 'Code', 'yasm': 'Code',
            'clj': 'Code', 'cljs': 'Code', 'cljc': 'Code',
            'edn': 'Code', 'ps1': 'Code', 'psm1': 'Code',
            'psd1': 'Code', 'ps1xml': 'Code', 'psc1': 'Code',
            'bat': 'Code', 'cmd': 'Code', 'com': 'Code',
            'exe': 'Code', 'msi': 'Code', 'app': 'Code',
            'appimage': 'Code', 'snap': 'Code', 'flatpak': 'Code',
            'makefile': 'Code', 'make': 'Code', 'cmake': 'Code',
            'dockerfile': 'Code', 'procfile': 'Code', 'rakefile': 'Code',
            'gemfile': 'Code', 'podfile': 'Code', 'vagrantfile': 'Code',
            'dockerignore': 'Config', 'gitignore': 'Config',
            'gitattributes': 'Config', 'gitmodules': 'Config',
            'gitkeep': 'Config', 'hgignore': 'Config',
            'bzrignore': 'Config', 'cvsignore': 'Config',
            'editorconfig': 'Config', 'eslintrc': 'Config',
            'prettierrc': 'Config', 'stylelintrc': 'Config',
            'babelrc': 'Config', 'tslint.json': 'Config',
            'pylintrc': 'Config', 'flake8': 'Config',
            'setup.py': 'Code', 'requirements.txt': 'Config',
            'pipfile': 'Config', 'poetry.lock': 'Config',
            'pyproject.toml': 'Config', 'tox.ini': 'Config',
            'pytest.ini': 'Config', 'conftest.py': 'Code',
            # 3D/CAD
            '3ds': '3D Model', 'blend': '3D Model', 'dwg': 'CAD',
            'dxf': 'CAD', 'fbx': '3D Model', 'obj': '3D Model',
            'stl': '3D Model', 'step': 'CAD', 'stp': 'CAD',
            'iges': 'CAD', 'igs': 'CAD', 'skp': '3D Model',
            'dae': '3D Model', 'kmz': '3D Model', 'kml': '3D Model',
            'abc': 'CAD', 'acis': 'CAD', 'sat': 'CAD',
            'sab': 'CAD', 'prt': 'CAD', 'asm': 'CAD',
            'catpart': 'CAD', 'catproduct': 'CAD', 'cgr': 'CAD',
            'ipt': 'CAD', 'iam': 'CAD', 'neu': 'CAD',
            'sldprt': 'CAD', 'sldasm': 'CAD', 'x_t': 'CAD',
            'x_b': 'CAD', 'jt': 'CAD', 'vda': 'CAD',
            'model': '3D Model', 'mesh': '3D Model', 'ma': '3D Model',
            'mb': '3D Model', 'max': '3D Model', 'lwo': '3D Model',
            'lws': '3D Model', 'lxo': '3D Model', 'lwo2': '3D Model',
            'lws2': '3D Model', 'c4d': '3D Model', 'cinema4d': '3D Model',
            'c4d': '3D Model', 'lwo': '3D Model', 'lws': '3D Model',
            'lxo': '3D Model', '3dm': '3D Model', 'rhino': '3D Model',
            'rws': '3D Model', 'r3d': '3D Model', 'bip': '3D Model',
            'bvh': '3D Model', 'csm': '3D Model', 'ase': '3D Model',
            'ogex': '3D Model', 'ply': '3D Model', 'pct': '3D Model',
            'pzm': '3D Model', 'ztl': '3D Model', 'zpr': '3D Model',
            'vpb': '3D Model', 'vpp': '3D Model', 'vpx': '3D Model',
            'x3d': '3D Model', 'wrl': '3D Model', 'vrml': '3D Model',
            'geo': '3D Model', 'pov': '3D Model', 'inc': '3D Model',
            # Fonts
            'ttf': 'Font', 'otf': 'Font', 'woff': 'Font',
            'woff2': 'Font', 'eot': 'Font', 'pfb': 'Font',
            'pfm': 'Font', 'afm': 'Font', 'fon': 'Font',
            'fnt': 'Font', 'suit': 'Font', 'dfont': 'Font',
            'psf': 'Font', 'pcf': 'Font', 'bdf': 'Font',
            'ttc': 'Font', 'font': 'Font', 'gai': 'Font',
            'sfd': 'Font', 'ufo': 'Font', 'tef': 'Font',
            # Credentials
            'cer': 'Credential', 'crt': 'Credential', 'p12': 'Credential',
            'pem': 'Credential', 'pfx': 'Credential', 'ppk': 'Credential',
            'ovpn': 'Credential', 'ssh': 'Credential', 'kdbx': 'Credential',
            'key': 'Credential', 'pub': 'Credential', 'asc': 'Credential',
            'gpg': 'Credential', 'sig': 'Credential', 'der': 'Credential',
            'csr': 'Credential', 'crl': 'Credential', 'p7b': 'Credential',
            'p7c': 'Credential', 'spc': 'Credential', 'sst': 'Credential',
            'stl': 'Credential', 'pkcs12': 'Credential', 'pkcs7': 'Credential',
            'jks': 'Credential', 'keystore': 'Credential', 'truststore': 'Credential',
            'ks': 'Credential', 'ts': 'Credential', 'jceks': 'Credential',
            'bks': 'Credential', 'ucf': 'Credential', 'kdb': 'Credential',
            'kdbx': 'Credential', 'agilekeychain': 'Credential',
            'keychain': 'Credential', 'wallet': 'Credential',
            'secret': 'Credential', 'env': 'Credential', 'env.local': 'Credential',
            'env.production': 'Credential', 'env.development': 'Credential',
            'env.staging': 'Credential', 'env.test': 'Credential',
            # Games
            'sav': 'Game', 'save': 'Game', 'dat': 'Game',
            'rom': 'Game', 'iso': 'Game', 'n64': 'Game',
            'z64': 'Game', 'v64': 'Game', 'nes': 'Game',
            'smc': 'Game', 'sfc': 'Game', 'gb': 'Game',
            'gbc': 'Game', 'gba': 'Game', 'nds': 'Game',
            '3ds': 'Game', 'cia': 'Game', '3dsx': 'Game',
            'nsp': 'Game', 'xci': 'Game', 'wad': 'Game',
            'wii': 'Game', 'wbfs': 'Game', 'dol': 'Game',
            'elf': 'Game', 'iso': 'Game', 'bin': 'Game',
            'cue': 'Game', 'ccd': 'Game', 'sub': 'Game',
            'mds': 'Game', 'mdf': 'Game', 'img': 'Game',
            'gcm': 'Game', 'dol': 'Game', 'elf': 'Game',
            'iso': 'Game', 'wbfs': 'Game', 'wad': 'Game',
            'nkit.iso': 'Game', 'rvz': 'Game', 'wia': 'Game',
            'nsp': 'Game', 'xci': 'Game', 'nca': 'Game',
            'ncz': 'Game', 'nro': 'Game', 'nacp': 'Game',
            'nrr': 'Game', 'nso': 'Game', 'nsz': 'Game',
            'xci': 'Game', 'xsz': 'Game', 'xcz': 'Game',
            'psv': 'Game', 'vpk': 'Game', 'pkg': 'Game',
            'rap': 'Game', 'rif': 'Game', 'edat': 'Game',
            'psarc': 'Game', 'pfs': 'Game', 'gp4': 'Game',
            'sfo': 'Game', 'sfm': 'Game', 'sfd': 'Game',
            'ps1': 'Game', 'ps2': 'Game', 'psx': 'Game',
            'psexe': 'Game', 'psf': 'Game', 'minipsf': 'Game',
            'cso': 'Game', 'zso': 'Game', 'iso': 'Game',
            'bin': 'Game', 'cue': 'Game', 'ccd': 'Game',
            'mds': 'Game', 'mdf': 'Game', 'img': 'Game',
            'nrg': 'Game', 'cdi': 'Game', 'gdi': 'Game',
            'ecm': 'Game', 'md5': 'Game', 'sha1': 'Game',
            'sha256': 'Game', 'sfv': 'Game', 'par2': 'Game',
            'par': 'Game', 'rev': 'Game', 'rar': 'Game',
            'zip': 'Game', '7z': 'Game',
        }
    
    def scan_and_categorize(self):
        """Scan folder and auto-categorize files in one step"""
        target_folder = self.folder_entry.text()
        if not target_folder:
            QMessageBox.warning(self, "Warning", "Please select a folder first.")
            return
        
        if not os.path.exists(target_folder):
            QMessageBox.warning(self, "Error", "Folder does not exist.")
            return
        
        # Create and start worker thread
        self.scan_worker = ScanWorker(target_folder, self.known_types)
        self.scan_worker.progress_updated.connect(self.progress_bar.setValue)
        self.scan_worker.status_updated.connect(self.status_label.setText)
        self.scan_worker.scan_complete.connect(self.on_scan_complete)
        self.scan_worker.start()
    
    def on_scan_complete(self, file_count):
        """Handle scan completion"""
        self.file_list = self.scan_worker.file_list
        self.progress_label.setText("Complete!")
        self.refresh_file_list()
        self.refresh_summary()
        self.update_category_filter()
    
    def update_category_filter(self):
        """Update category filter dropdown"""
        categories = ["All"]
        if self.file_list:
            categories.extend(sorted(set(f['category'] for f in self.file_list if f['category'])))
        self.category_filter.clear()
        self.category_filter.addItems(categories)
    
    def cancel_operation(self):
        """Cancel current operation"""
        if hasattr(self, 'scan_worker') and self.scan_worker.isRunning():
            self.scan_worker.cancel()
            self.progress_label.setText("Cancelling...")
            self.status_label.setText("Operation cancelled by user")


if __name__ == "__main__":
    app = QApplication([])
    window = FileWhipApp()
    window.show()
    app.exec()
